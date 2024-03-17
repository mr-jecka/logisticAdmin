import psycopg2
import openpyxl
import uuid
import os
import logging
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine, func,  Column, Integer, String, Numeric, Boolean, Time, Date, and_
from sqlalchemy.dialects.postgresql import UUID
from sqlalchemy.ext.declarative import declarative_base
from datetime import timedelta
import requests


Base = declarative_base()


logging.basicConfig(level=logging.INFO)
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

user = os.getenv('POSTGRES_USER', 'bank_user')
password = os.getenv('POSTGRES_PASSWORD', 'bank_password')
host = os.getenv('POSTGRES_HOST', 'localhost')
#host = os.getenv('POSTGRES_HOST', '194.87.92.15')
port = os.getenv('POSTGRES_PORT_NUMBER', "5432")
database = os.getenv('POSTGRES_DB', 'bank')

conn = {
    'user': user,
    'password': password,
    'host': host,
    'port': port,
    'database': database
}

connection_string = f'postgresql://{user}:{password}@{host}:{port}/{database}'
engine = create_engine(connection_string)
Session = sessionmaker(bind=engine)
session = Session()


class DriversTable(Base):
    __tablename__ = 'drivers'

    id = Column(Integer, primary_key=True)
    num_car = Column(String(255), nullable=False)
    last_name = Column(String(255), nullable=False)
    first_name = Column(String(255), nullable=False)
    patronymic = Column(String(255))
    user_id = Column(Integer)
    location = Column(Integer)


class ReestrTable(Base):
    __tablename__ = 'reestr_table'

    num_th = Column(String(20))
    date_th = Column(Date)
    total_count_boxes = Column(Integer)
    total_weight = Column(Numeric(10, 2))
    lastname = Column(String(255))
    notification = Column(Boolean)
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    num_car = Column(String(255))
    shipment_time = Column(Time)
    arrival_time = Column(Time)
    departure_time = Column(Time)
    arrival_time_fact = Column(Time)
    user_id = Column(Integer)
    location = Column(String(255))
    sum_points = Column(Integer)


class AddressTable(Base):
    __tablename__ = 'address_table'

    num_route = Column(String)
    num_shop = Column(String(20))
    code_tt = Column(String(20))
    address_delivery = Column(String(255))
    count_boxes = Column(Integer)
    weight = Column(Numeric(10, 2))
    th_id = Column(Integer)
    num_th = Column(String(20))
    last_name = Column(String(255))
    arrival_time = Column(Time)
    id = Column(UUID, primary_key=True)
    departure_time = Column(Time)
    index_number = Column(Integer)
    user_id = Column(Integer)
    longitude = Column(Numeric)
    latitude = Column(Numeric)
    date_th = Column(Date)
    location = Column(String(255))
    priority = Column(Integer)


def calculate_delivery_order(departure_place, api_key, destinations):
    finish_road = []
    while any(not v[1] for v in destinations.values()):
        dest = "|".join(k for k, v in destinations.items() if not v[1])
        params = {
            "origins": departure_place,
            "destinations": dest,
            "mode": "truck",
            "apikey": api_key
        }
        res = requests.get("https://api.routing.yandex.net/v2/distancematrix", params=params).json()
        rows = res["rows"][0]['elements']
        while True:
            try:
                rows.remove({'status': 'FAIL'})
            except:
                break
        result_dest = [{"coord": coord, "distance": row['distance']['value']} for coord, row in zip(dest.split('|'), rows)]
        sorted_rows = sorted(result_dest, key=lambda d: d['distance'])
        if len(sorted_rows)>0:
            closest_coord = sorted_rows[0]['coord']
            finish_road.append(closest_coord)
            destinations[closest_coord][1] = True
            departure_place = closest_coord
        else:
            break
    return finish_road


# def update_index_numbers(day):
#     connection = psycopg2.connect(**conn)
#     cursor = connection.cursor()
#     cursor.execute("SELECT num_th FROM public.reestr_table WHERE user_id IS NULL and date_th = %s", (day,))
#     num_ths = cursor.fetchall()
#
#     for num_th in num_ths:
#         cursor.execute("SELECT latitude, longitude, address_delivery FROM public.address_table WHERE num_th = %s", (num_th,))
#         destinations_data = cursor.fetchall()
#         destinations = {f"{data[0]},{data[1]}": [data[2], False] for data in destinations_data}
#
#         ordered_addresses = calculate_delivery_order("55.716498,36.841024", "a6f0bc4e-a4ed-4839-8a91-25c403ff4b46", destinations)
#         for index, coord in enumerate(ordered_addresses, start=1):
#             address = destinations[coord][0]
#             cursor.execute("UPDATE public.address_table SET index_number = %s WHERE num_th = %s AND address_delivery = %s", (index, num_th, address))
#     connection.commit()


def update_index_numbers():
    try:
        today = datetime.now().date()
        num_th_list = (
            session.query(AddressTable.num_th)
            .filter(AddressTable.date_th == today)
            .distinct()
            .all()
        )

        for num_th, in num_th_list:
            address_ids = (
                session.query(AddressTable.id)
                .filter(AddressTable.date_th == today, AddressTable.num_th == num_th)
                .order_by(AddressTable.location, AddressTable.priority)
                .all()
            )

            for index, (address_id,) in enumerate(address_ids, start=1):
                session.query(AddressTable).filter(AddressTable.id == address_id).update({"index_number": index})

        session.commit()
    except Exception as e:
        print(f"Database error: {e}")
        session.rollback()


def get_routes_and_insert_index(day):
    connection = psycopg2.connect(**conn)
    try:
        cursor = connection.cursor()
        query_update_index = """
        WITH sorted_addresses AS (
            SELECT id, ROW_NUMBER() OVER(PARTITION BY num_th ORDER BY id) AS rn
            FROM public.address_table
            WHERE num_th IN (SELECT num_th FROM public.reestr_table WHERE user_id IS NULL and date_th = %s)
        )
        UPDATE public.address_table
        SET index_number = sorted_addresses.rn
        FROM sorted_addresses
        WHERE public.address_table.id = sorted_addresses.id;
        """
        cursor.execute(query_update_index, (day,))
        connection.commit()
    except psycopg2.Error as e:
        print(f"Database error: {e}")
        return []
    finally:
        connection.close()


def get_routes(day):
    connection = psycopg2.connect(**conn)
    try:
        cursor = connection.cursor()
        query = "SELECT num_th FROM public.reestr_table WHERE user_id IS NULL and date_th = %s"
        cursor.execute(query, (day,))

        routes = cursor.fetchall()
        return routes
    except (Exception, psycopg2.Error) as error:
        print("Error fetching routes from the database:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()


def get_info_for_report():
    connection = psycopg2.connect(**conn)
    today = datetime.now().date()
    try:
        cursor = connection.cursor()
        cursor.execute("SELECT num_th, lastname, num_car, arrival_time_fact, shipment_time, departure_time, arrival_time FROM public.reestr_table WHERE date_th = %s",
                       (today,))
        routes = cursor.fetchall()
        return routes
    except (Exception, psycopg2.Error) as error:
        print("Error fetching routes from the database:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()


def get_driver_last_name(driver_id):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        cursor.execute(f"SELECT last_name FROM public.drivers WHERE car_number = '{driver_id}'")
        last_name = cursor.fetchone()
        if last_name:
            return last_name[0]
        else:
            return None
    except (Exception, psycopg2.Error) as error:
        print("Error fetching driver's last name from the database:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()


def insert_user_id_for_addresses(selected_route, user_id):
    connection = None
    cursor = None
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        update_query = """
        UPDATE public.address_table
        SET user_id = %s
        WHERE num_th = %s AND user_id IS NULL;
        """
        cursor.execute(update_query, (user_id, selected_route))
        updated_rows = cursor.rowcount  # Количество обновленных строк
        connection.commit()
        if updated_rows > 0:
            logger.info(f"User ID {user_id} присвоен для маршрута {selected_route}. Обновлено записей: {updated_rows}")
        else:
            logger.warning(f"Нет записей для обновления user_id для маршрута {selected_route}")
    except Exception as e:
        logger.exception("Ошибка при обновлении user_id в адресной таблице:")
        if connection:
            connection.rollback()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


def insert_driver_for_route(selected_route, selected_driver, selected_driver_car, loading_time, user_id):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        cursor.execute(
            f"UPDATE public.reestr_table SET lastname = %s, num_car = %s, arrival_time = %s, user_id = %s WHERE num_th = %s",
            (selected_driver, selected_driver_car, loading_time, user_id, selected_route)
        )
        connection.commit()
    except (Exception, psycopg2.Error) as error:
        print("Error updating reestr_table in the database:", error)
        connection.rollback()
    finally:
        if connection:
            cursor.close()
            connection.close()


from decimal import Decimal


def convert_decimal(value):  # Конвертируем Decimal в float
    if isinstance(value, Decimal):
        return float(value)
    return value


def filter_addresses(session, num_th):
    addresses = (
        session.query(AddressTable)
        .filter_by(num_th=num_th)
        .all()
    )
    return addresses


def get_optimal_json():
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)

    try:
        #reestr_entries = session.query(AddressTable).filter(func.date(AddressTable.date_th) == today).all()
        reestr_entries = session.query(AddressTable.num_th).filter(func.date(AddressTable.date_th) == today).distinct().all()
        # last_names = session.query(AddressTable.last_name).filter(func.date(AddressTable.date_th) == today).distinct().all()

        route_entries = []
        addresses_count = {}

        for entry in reestr_entries:
            num_th = entry.num_th
            addresses = filter_addresses(session, num_th)

            address_list = []

            driver = session.query(AddressTable.last_name).filter(AddressTable.num_th == num_th).first()
            driver_name = driver.last_name if driver else None

            if driver_name is None:
                logging.warning(f"Фамилия водителя в маршруте {num_th} равна None")

            for addr in addresses:
                weight = convert_decimal(addr.weight)

                address_info = {
                    "num_route": addr.num_route,
                    "num_shop": addr.num_shop,
                    "code_tt": addr.code_tt,
                    "address_delivery": addr.address_delivery,
                    "index_number": addr.index_number,
                    "count_boxes": convert_decimal(addr.count_boxes),
                    "weight": weight
                }

                address_list.append(address_info)

            address_list.sort(key=lambda x: x['index_number'])

            addresses_count[num_th] = len(address_list)

            route_entry = {
                "num_th": num_th,
                "date_th": today.strftime('%d.%m.%Y'),
                "driver": driver_name,
                "num_car": None,
                "addresses": address_list
            }

            route_entries.append(route_entry)

        return route_entries, addresses_count

    except Exception as e:
        logging.error(f"Произошла ошибка: {e}")
        logging.exception("Ошибка во время выполнения кода:")
        return []


# def get_optimal_json():
#     today = datetime.now().date()
#     tomorrow = today + timedelta(days=1)
#     drivers_shops = get_likes_shop_drivers()
#
#     try:
#         reestr_entries = session.query(ReestrTable).filter(func.date(ReestrTable.date_th) == tomorrow).all()
#
#         route_entries = []
#         overweight_addresses = []
#
#         for entry in reestr_entries:
#             num_th = entry.num_th
#             addresses = filter_addresses(session, num_th)
#
#             address_list = []
#             total_weight = 0
#             driver_assigned = None
#             num_car = None
#
#             for addr in addresses:
#                 weight = convert_decimal(addr.weight)
#                 total_weight += weight
#
#                 address_info = {
#                     "num_route": addr.num_route,
#                     "num_shop": addr.num_shop,
#                     "code_tt": addr.code_tt,
#                     "address_delivery": addr.address_delivery,
#                     "index_number": addr.index_number,
#                     "count_boxes": convert_decimal(addr.count_boxes),
#                     "weight": weight
#                 }
#
#                 address_list.append(address_info)
#
#                 if total_weight > 3000:
#                     overweight_addresses.append(address_info)
#
#                 if not driver_assigned and drivers_shops.get(addr.num_shop):
#                     driver_assigned, num_car = drivers_shops.get(addr.num_shop)
#
#             route_entries.append({
#                 "num_th": num_th,
#                 "date_th": entry.date_th.strftime('%d.%m.%Y'),
#                 "driver": driver_assigned,
#                 "num_car": num_car,
#                 "addresses": address_list
#                 })
#
#         return route_entries, overweight_addresses
#
#     except Exception as e:
#         print("An error occurred:", e)
#         return []


def get_likes_shop_drivers():
    drivers = session.query(DriversTable).all()
    drivers_likes = {}
    for driver in drivers:
        if driver.location:
            liked_shops = [location.strip() for location in driver.location.split(',')]
            drivers_likes[driver.last_name] = liked_shops
    logging.info(f"Магазины водителей: {drivers_likes}")
    return drivers_likes


def assign_driver_to_th(th, drivers_shops):
    shops_in_th = {addr["num_shop"] for addr in th["addresses"]}
    for driver, liked_shops in drivers_shops.items():
        if any(shop in liked_shops for shop in shops_in_th):
            for addr in th["addresses"]:
                addr["driver_last_name"] = driver
            logging.info(f"Assigned driver {driver} to num_th {th['num_th']} with shops {shops_in_th}")
            return
    logging.warning(f"No suitable driver found for num_th {th['num_th']} with shops {shops_in_th}")



# def get_likes_shop_drivers():
#     drivers = session.query(DriversTable).all()
#     like_shop = {}
#     for driver in drivers:
#         if driver.like_num_shop:
#             for num_shop in driver.like_num_shop.split(','):
#                 like_shop[num_shop.strip()] = (driver.last_name, driver.num_car)
#
#     return like_shop


# def get_optimal_json():
#     today = datetime.now().date()
#     tomorrow = today + timedelta(days=1)
#     try:
#         reestr_entries = session.query(ReestrTable).filter(func.date(ReestrTable.date_th) == tomorrow).all()
#
#         route_entries = []
#         overweight_addresses = []
#
#         for entry in reestr_entries:
#             num_th = entry.num_th
#             addresses = filter_addresses(session, num_th)
#
#             address_list = []
#             total_weight = 0
#
#             for addr in addresses:
#                 weight = convert_decimal(addr.weight)
#                 total_weight += weight
#
#                 if total_weight > 3000:
#                     overweight_addresses.append({
#                         "num_route": addr.num_route,
#                         "num_shop": addr.num_shop,
#                         "code_tt": addr.code_tt,
#                         "address_delivery": addr.address_delivery,
#                         "index_number": addr.index_number,
#                         "count_boxes": convert_decimal(addr.count_boxes),
#                         "weight": weight
#                     })
#                 else:
#                     address_list.append({
#                         "num_route": addr.num_route,
#                         "num_shop": addr.num_shop,
#                         "code_tt": addr.code_tt,
#                         "address_delivery": addr.address_delivery,
#                         "index_number": addr.index_number,
#                         "count_boxes": convert_decimal(addr.count_boxes),
#                         "weight": weight
#                     })
#
#         route_entries.append({
#             "num_th": num_th,
#             "date_th": entry.date_th.strftime('%d.%m.%Y'),
#             "addresses": address_list
#         })
#
#         final_data = route_entries + [{"overweight_addresses": overweight_addresses}]
#
#         return final_data
#
#     except Exception as e:
#         print("An error occurred:", e)
#         return []


# def get_optimal_json():
#     today = datetime.now().date()
#     tomorrow = today + timedelta(days=1)
#     try:
#         reestr_entries = session.query(ReestrTable).filter(func.date(ReestrTable.date_th) == tomorrow).all()
#         result = []
#
#         for entry in reestr_entries:
#             print("Processing num_th:", entry.num_th)
#             num_th = entry.num_th
#             addresses = filter_addresses(session, num_th)
#
#             address_list = []
#             for idx, addr in enumerate(addresses):
#                 address_list.append({
#                     "index_number": addr.index_number,
#                     "num_route": addr.num_route,
#                     "num_shop": addr.num_shop,
#                     "code_tt": addr.code_tt,
#                     "address_delivery": addr.address_delivery,
#                     "count_boxes": convert_decimal(addr.count_boxes),
#                     "weight": convert_decimal(addr.weight)
#                 })
#
#             result.append({
#                 "num_th": num_th,
#                 "date_th": entry.date_th.strftime('%d.%m.%Y'),
#                 "addresses": address_list
#             })
#
#         return result
#     except Exception as e:
#         print("An error occurred:", e)
#         return []


def fetch_addresses(session, num_th, is_null):
    condition = AddressTable.arrival_time.is_(None) if is_null else AddressTable.arrival_time.isnot(None)
    addresses = (
        session.query(AddressTable)
        .filter_by(num_th=num_th)
        .filter(condition)
        .order_by(AddressTable.arrival_time.asc())
        .all()
    )
    return addresses


def get_main_json():
    today = datetime.now().date()

    try:
        reestr_entries = session.query(ReestrTable).filter(func.date(ReestrTable.date_th) == today).all()
        result = []

        for entry in reestr_entries:
            num_th = entry.num_th
            addresses = fetch_addresses(session, num_th, False)
            addresses += fetch_addresses(session, num_th, True)

            for addr in addresses:
                if addr.num_th is not None:
                    result.append({
                        "num_route": addr.num_route,
                        "arrival_time": addr.arrival_time,
                        "departure_time": addr.departure_time
                    })

        return result

    except Exception as e:
        print(f"Error: {e}")
        return []


from datetime import datetime, time


def fetch_addresses_14(session, num_th, is_null):
    condition = AddressTable.arrival_time.is_(None) if is_null else AddressTable.arrival_time.isnot(None)
    time_condition = AddressTable.arrival_time < time(14, 0, 0)
    addresses = (
        session.query(AddressTable)
        .filter_by(num_th=num_th)
        .filter(condition)
        .filter(time_condition)
        .order_by(AddressTable.arrival_time.asc())
        .all()
    )
    return addresses


def get_main_json_14():
    today = datetime.now().date()
    try:
        reestr_entries = session.query(ReestrTable).filter(func.date(ReestrTable.date_th) == today).all()
        result = []

        for entry in reestr_entries:
            num_th = entry.num_th
            addresses = fetch_addresses_14(session, num_th, False)
            addresses += fetch_addresses_14(session, num_th, True)

            for addr in addresses:
                if addr.num_th is not None:
                    result.append(addr.num_route)
        return result

    except Exception as e:
        print(f"Error: {e}")
        return []


def update_driver_assignment(driver_id, selected_route):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        cursor.execute(f"UPDATE public.drivers SET user_id = {selected_route} WHERE car_number = '{driver_id}'")
        connection.commit()
    except (Exception, psycopg2.Error) as error:
        print("Error updating driver assignment in the database:", error)
        connection.rollback()
    finally:
        if connection:
            cursor.close()
            connection.close()


def get_driver_last_names():
    try:
        last_names = session.query(AddressTable.last_name).distinct().all()

        last_name_list = [entry.last_name for entry in last_names]
        return last_name_list

    except Exception as e:
        print(f"Error: {e}")
        return []


def num_th_to_drivers():
    num_th_list = get_num_th_and_drivers()
    driver_last_names = get_driver_last_names()
    if len(num_th_list) < len(driver_last_names):
        num_th_to_add = len(driver_last_names) - len(num_th_list)
        max_num_th = max([int(num_th.split('-')[1]) for num_th in num_th_list], default=0)

        for _ in range(num_th_to_add):
            max_num_th += 1
            new_num_th = f"0000-{max_num_th:06d}"
            num_th_list.append(new_num_th)

    print("Список num_th:")
    print(num_th_list)

    print("Список фамилий водителей:")
    print(driver_last_names)

    for i in range(len(driver_last_names)):
        driver_last_name = driver_last_names[i]
        if i < len(num_th_list):
            num_th = num_th_list[i]
            session.query(AddressTable).filter(AddressTable.last_name == driver_last_name).update({"num_th": num_th})

    session.commit()


def get_num_th_and_drivers():
    today = datetime.now().date()
    try:
        num_ths = session.query(ReestrTable.num_th).filter(func.date(ReestrTable.date_th) == today).all()

        num_th_list = [entry.num_th for entry in num_ths]

        return num_th_list

    except Exception as e:
        print(f"Error: {e}")
        return []


def get_drivers_for_route():
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        cursor.execute("SELECT last_name, num_car, user_id FROM public.drivers WHERE user_id IS NOT NULL")
        drivers = cursor.fetchall()
        return drivers
    except (Exception, psycopg2.Error) as error:
        print("Error fetching available drivers from the database:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()


def generate_and_assign_uuid(cursor, table_name, id_column):
    new_uuid = uuid.uuid4()
    new_uuid_str = new_uuid.hex  # Преобразование UUID в строку
    cursor.execute(f"""
        UPDATE {table_name}
        SET {id_column} = %s
        WHERE {id_column} IS NULL;
    """, (new_uuid_str,))
    cursor.connection.commit()
    return new_uuid


def find_and_delete_total(sheet, start_row):
    total_row_index = None
    for row_index, row in enumerate(
            sheet.iter_rows(min_row=start_row, max_col=1, max_row=sheet.max_row, values_only=True), start=start_row):
        if row[0] == "Итого":
            total_row_index = row_index
            break

    if total_row_index is not None:
        sheet.delete_rows(total_row_index, 21)  # Удалить строку "Итого" и следующие 20 строк


def insert_excel_to_db(excel_file_path, db_params):
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
    except psycopg2.Error as e:
        logging.error(f"Error connecting to the database: {e}")
        return
    try:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = wb.active

        find_and_delete_total(sheet, start_row=5)


        ths = []
        th = {}

        for row in sheet.iter_rows(min_row=5, values_only=True):
            if row[2] is not None:
                logging.info(f"Processing row: {row}")
                if len(th.keys()) != 0:
                    if th.get("total_weight", 0) >= 1000:  # Проверяем вес предыдущего num_th
                        ths.append(th)
                th = {}
                th["num_th"] = row[2]
                date_str = row[3]
                if isinstance(date_str, datetime):
                    date_str = date_str.strftime('%d.%m.%Y')
                try:
                    date_th = datetime.strptime(date_str, '%d.%m.%Y')
                    th["date_th"] = date_th.strftime('%Y-%m-%d')
                except ValueError:
                    logging.error(f"Error converting date string: {date_str}")
                th["total_count_boxes"] = int(row[9])
                th["total_weight"] = float(row[10])
                th["addresses"] = []
            else:
                addr = {}
                addr["num_th"] = th["num_th"]
                addr["date_th"] = th["date_th"]
                addr["num_route"] = row[4]
                addr["num_shop"] = row[6]
                addr["code_tt"] = row[7]
                addr["address_delivery"] = row[8]
                addr["count_boxes"] = int(row[9])
                addr["weight"] = float(row[10])
                th["addresses"].append(addr)
        if th.get("total_weight", 0) >= 1000:
            ths.append(th)

        with conn:
            with conn.cursor() as cursor:
                for th in ths:
                    cursor.execute("""
                        INSERT INTO reestr_table (num_th, date_th, total_count_boxes, total_weight)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id;
                    """, (th["num_th"], th["date_th"], th["total_count_boxes"], th["total_weight"]))
                    th_id = cursor.fetchone()[0]
                    logging.info(
                        f"Recorded in reestr_table: num_th={th['num_th']}")  # Логирование записи в reestr_table

                    for addr in th["addresses"]:
                        cursor.execute("""
                            SELECT latitude, longitude, location, priority FROM public.addresses WHERE code_tt = %s;
                        """, (addr["code_tt"],))
                        result = cursor.fetchone()
                        if result:
                            addr["latitude"], addr["longitude"], addr["location"], addr["priority"] = result

                            cursor.execute("""
                                INSERT INTO address_table (
                                num_th, date_th, num_route, num_shop, code_tt, address_delivery,
                                 count_boxes, weight, th_id, latitude, longitude, location, priority)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                            """, (addr["num_th"], addr["date_th"], addr["num_route"], addr["num_shop"], addr["code_tt"],
                                  addr["address_delivery"], addr["count_boxes"], addr["weight"],
                                  th_id, addr["latitude"], addr["longitude"], addr["location"], addr["priority"]))
                            logging.info(
                                f"Recorded in address_table: num_th={addr['num_th']}")  # Логирование записи в address_table

                            new_uuid_in_reestr = generate_and_assign_uuid(cursor, 'reestr_table', 'id')
                            new_uuid_in_address = generate_and_assign_uuid(cursor, 'address_table', 'id')
                            print("New UUID in reestr_table:", new_uuid_in_reestr)
                            print("New UUID in address_table:", new_uuid_in_address)

        logging.info("Data inserted into the database successfully")
        return True, "Data inserted into the database successfully"
    except Exception as e:
        logging.error(f"Error inserting data into the database: {e}")
        return False, f"Error inserting data into the database: {e}"
    finally:
        conn.close()


# def insert_excel_to_db(excel_file_path, db_params):
#     try:
#         conn = psycopg2.connect(**db_params)
#         cursor = conn.cursor()
#     except psycopg2.Error as e:
#         logging.error(f"Error connecting to the database: {e}")
#         return
#     try:
#         wb = openpyxl.load_workbook(excel_file_path, data_only=True)
#         sheet = wb.active
#
#         find_and_delete_total(sheet, start_row=5)
#
#
#         ths = []
#         th = {}
#
#         for row in sheet.iter_rows(min_row=5, values_only=True):
#             if row[2] is not None:
#                 logging.info(f"Processing row: {row}")
#                 if len(th.keys()) != 0:
#                     if th.get("total_weight", 0) >= 250:  # Проверяем вес предыдущего num_th
#                         ths.append(th)
#                 th = {}
#                 th["num_th"] = row[2]
#                 date_str = row[3]
#                 if isinstance(date_str, datetime):
#                     date_str = date_str.strftime('%d.%m.%Y')
#                 try:
#                     date_th = datetime.strptime(date_str, '%d.%m.%Y')
#                     th["date_th"] = date_th.strftime('%Y-%m-%d')
#                 except ValueError:
#                     logging.error(f"Error converting date string: {date_str}")
#                 th["total_count_boxes"] = int(row[9])
#                 th["total_weight"] = float(row[10])
#                 th["addresses"] = []
#             else:
#                 addr = {}
#                 addr["num_th"] = th["num_th"]
#                 addr["date_th"] = th["date_th"]
#                 addr["num_route"] = row[4]
#                 addr["num_shop"] = row[6]
#                 addr["code_tt"] = row[7]
#                 addr["address_delivery"] = row[8]
#                 addr["count_boxes"] = int(row[9])
#                 addr["weight"] = float(row[10])
#                 th["addresses"].append(addr)
#         ths.append(th)
#
#         with conn:
#             with conn.cursor() as cursor:
#                 for th in ths:
#                     cursor.execute("""
#                         INSERT INTO reestr_table (num_th, date_th, total_count_boxes, total_weight)
#                         VALUES (%s, %s, %s, %s)
#                         RETURNING id;
#                     """, (th["num_th"], th["date_th"], th["total_count_boxes"], th["total_weight"]))
#                     th_id = cursor.fetchone()[0]
#
#                     for addr in th["addresses"]:
#                         coordinates = revers_geocoding_yandex(addr["address_delivery"])
#                         if coordinates:
#                             addr["latitude"], addr["longitude"] = coordinates
#                             cursor.execute("""
#                                 INSERT INTO address_table (
#                                 num_th, num_route, num_shop, code_tt, address_delivery,
#                                  count_boxes, weight, th_id, latitude, longitude)
#                                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
#                             """, (addr["num_th"], addr["num_route"], addr["num_shop"], addr["code_tt"],
#                                   addr["address_delivery"], addr["count_boxes"], addr["weight"],
#                                   th_id, addr["latitude"], addr["longitude"]))
#                             new_uuid_in_reestr = generate_and_assign_uuid(cursor, 'reestr_table', 'id')
#                             new_uuid_in_address = generate_and_assign_uuid(cursor, 'address_table', 'id')
#                             print("New UUID in reestr_table:", new_uuid_in_reestr)
#                             print("New UUID in address_table:", new_uuid_in_address)
#
#         logging.info("Data inserted into the database successfully")
#         return True, "Data inserted into the database successfully"
#     except Exception as e:
#         logging.error(f"Error inserting data into the database: {e}")
#         return False, f"Error inserting data into the database: {e}"
#     finally:
#         conn.close()


def revers_geocoding_yandex(address):
    token = '5d9f623c-b003-4632-956f-464db92caaae'
    headers = {"Accept-Language": "ru"}
    response = requests.get(
        f'https://geocode-maps.yandex.ru/1.x/?apikey={token}&geocode={address}&format=json',
        headers=headers).json()
    if (
            "response" in response
            and "GeoObjectCollection" in response["response"]
            and "featureMember" in response["response"]["GeoObjectCollection"]
    ):
        geo_objects = response["response"]["GeoObjectCollection"]["featureMember"]
        if geo_objects:
            first_result = geo_objects[0]["GeoObject"]
            latitude, longitude = map(float, first_result["Point"]["pos"].split())
            return latitude, longitude
    return None


def insert_scheduled_arrival(user_id, date_time_obj):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()
        operation_uuid = str(uuid.uuid4())
        current_date = datetime.now().date()

        insert_query = """
            INSERT INTO reports (id, user_id, date, scheduled_arrival, car_number, last_name, first_name, patronymic)
            SELECT %s, %s, %s, %s, d.car_number, d.last_name, d.first_name, d.patronymic
            FROM drivers AS d
            WHERE d.user_id = %s
            ON CONFLICT DO NOTHING
        """
        cursor.execute(insert_query, (operation_uuid, user_id, current_date, date_time_obj, user_id))

        connection.commit()
        cursor.close()
        connection.close()
        print("Data inserted successfully into the 'reports' table.")
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to the database or inserting data:", error)


def insert_actual_arrival(user_id, date_time_obj):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()

        current_date = datetime.now().date()
        select_query = "SELECT * FROM reports WHERE user_id = %s AND date = %s"
        cursor.execute(select_query, (user_id, current_date))
        existing_record = cursor.fetchone()

        if existing_record:
            update_query = "UPDATE reports SET actual_arrival = %s WHERE user_id = %s AND date = %s"
            cursor.execute(update_query, (date_time_obj, user_id, current_date))
        else:
            operation_uuid = str(uuid.uuid4())
            insert_query = "INSERT INTO reports (id, user_id, date, actual_arrival) VALUES (%s, %s, %s, %s)"
            cursor.execute(insert_query, (operation_uuid, user_id, current_date, date_time_obj))

        connection.commit()
        cursor.close()
        connection.close()
        print("Data inserted successfully into the 'reports' table")
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to the database or inserting data:", error)


def insert_shipment(user_id, date_time_obj):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()

        current_date = datetime.now().date()
        select_query = "SELECT * FROM reports WHERE user_id = %s AND date = %s AND scheduled_arrival IS NOT NULL"
        cursor.execute(select_query, (user_id, current_date))
        existing_record = cursor.fetchone()

        if existing_record:
            update_query = "UPDATE reports SET shipment = %s WHERE user_id = %s AND date = %s"
            cursor.execute(update_query, (date_time_obj, user_id, current_date))
        else:
            operation_uuid = str(uuid.uuid4())
            insert_query = "INSERT INTO reports (id, user_id, date, shipment) VALUES (%s, %s, %s, %s)"
            cursor.execute(insert_query, (operation_uuid, user_id, current_date, date_time_obj))

        connection.commit()
        cursor.close()
        connection.close()
        print("Data inserted successfully into the 'report' table")
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to the database or inserting data:", error)


def insert_departure(user_id, date_time_obj):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()

        current_date = datetime.now().date()
        select_query = "SELECT * FROM reports WHERE user_id = %s AND date = %s AND scheduled_arrival IS NOT NULL"
        cursor.execute(select_query, (user_id, current_date))
        existing_record = cursor.fetchone()

        if existing_record:
            update_query = "UPDATE reports SET departure = %s WHERE user_id = %s AND date = %s"
            cursor.execute(update_query, (date_time_obj, user_id, current_date))
        else:
            operation_uuid = str(uuid.uuid4())
            insert_query = "INSERT INTO reports (id, user_id, date, departure) VALUES (%s, %s, %s, %s)"
            cursor.execute(insert_query, (operation_uuid, user_id, current_date, date_time_obj))

        connection.commit()
        cursor.close()
        connection.close()
        print("Data inserted successfully into the 'report' table")
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to the database or inserting data:", error)


def get_internal_report(report_date):
    try:
        connection = psycopg2.connect(**conn)
        cursor = connection.cursor()

        select_query = "SELECT car_number, last_name, first_name, patronymic, scheduled_arrival, actual_arrival, shipment, departure FROM public.reports WHERE date = %s"
        cursor.execute(select_query, (report_date,))
        daily_report = cursor.fetchall()

        cursor.close()
        connection.close()

        return daily_report
    except (Exception, psycopg2.Error) as error:
        print("Error while connecting to the database or executing query:", error)


from datetime import datetime, timedelta
from collections import defaultdict
import math


def calculate_weights_and_get_driver_last_names():
    try:
        total_weight_tomorrow = 0.0
        weights_by_location = defaultdict(float)
        tomorrow = datetime.now().date() + timedelta(days=1)
        today = datetime.now().date()
        details_by_code_tt = []

        addresses_tomorrow = session.query(AddressTable).filter(AddressTable.date_th == today).all()

        for address in addresses_tomorrow:
            address_weight = float(address.weight.real)
            total_weight_tomorrow += address_weight

            first_digit = next((char for char in address.location if char.isdigit()), None)
            if first_digit:
                rounded_weight = math.ceil(address_weight)
                weights_by_location[first_digit] += rounded_weight

            detail = {
                'code_tt': address.code_tt,
                'weight': address.weight,
                'location': address.location,
                'priority': address.priority
            }
            details_by_code_tt.append(detail)

        total_batches = math.ceil(total_weight_tomorrow / 3000)
        logger.info(f"Total weight for tomorrow: {total_weight_tomorrow} kg, requiring {total_batches} drivers")

        for location, weight in weights_by_location.items():
            logger.info(f"Total weight for location {location}: {int(weight)} kg")

        return details_by_code_tt
    except Exception as e:
        logger.error(f"Error in calculate_weights_and_get_driver_last_names: {e}")
        return None


def actually_weight_reestr_table(total_weights_by_num_th):
    try:
        weights_dict = {num_th: data['total_weight'] for num_th, data in total_weights_by_num_th}

        query = session.query(ReestrTable).filter(ReestrTable.num_th.in_(weights_dict.keys()))

        for record in query:
            record.total_weight = weights_dict[record.num_th]
        session.commit()
        return "Weights updated successfully"

    except Exception as e:
        print(f"Error: {e}")
        session.rollback()
        return None


def actually_details_num_th():
    try:
        query = session.query(
            ReestrTable.num_th,
            ReestrTable.total_weight,
        )

        results = query.all()

        #actually_details_num_th = {}

        for num_th, total_weight in results:
            actually_details_num_th[num_th] = total_weight

        return actually_details_num_th

    except Exception as e:
        print(f"Error: {e}")
        return None


def calculate_sum_points(details):
    num_code_tt = len({entry['code_tt'] for entry in details if entry['code_tt'] is not None})
    return num_code_tt


def calculate_details_by_num_th():
    try:
        existing_none_row = session.query(ReestrTable).filter(ReestrTable.num_th == None).first()

        query = session.query(AddressTable.num_th,
                              AddressTable.location,
                              AddressTable.priority,
                              AddressTable.weight,
                              AddressTable.count_boxes,
                              AddressTable.code_tt)

        results = query.all()

        details_by_num_th = {}
        total_weights_by_num_th = {}

        for num_th, location, priority, weight, count_boxes, code_tt in results:
            if num_th not in details_by_num_th:
                details_by_num_th[num_th] = []

            details_by_num_th[num_th].append({
                'location': location,
                'priority': priority,
                'weight': weight,
                'count_boxes': count_boxes,
                'code_tt': code_tt})

            if num_th not in total_weights_by_num_th:
                total_weights_by_num_th[num_th] = {
                    'total_weight': 0, 'total_count_boxes': 0, 'locations': set(), 'sum_points': 0}

            total_weights_by_num_th[num_th]['total_weight'] += weight
            total_weights_by_num_th[num_th]['total_count_boxes'] += count_boxes
            total_weights_by_num_th[num_th]['locations'].add(location)

        for num_th, details in details_by_num_th.items():
            total_weights_by_num_th[num_th]['sum_points'] = calculate_sum_points(details)

        sorted_total_weights = sorted(
            total_weights_by_num_th.items(), key=lambda item: item[1]['total_weight'], reverse=True)

        print(sorted_total_weights)
        for num_th, data in sorted_total_weights:
            total_weight = data['total_weight']
            total_count_boxes = data['total_count_boxes']
            locations = ', '.join(str(location) for location in data['locations'])
            sum_points = data['sum_points']

            logging.debug(
                f"num_th: {num_th}, total_weight: {total_weight}, locations: {locations}, sum_points: {sum_points}")

            if num_th is None and existing_none_row:
                existing_none_row.total_weight = total_weight
                existing_none_row.total_count_boxes = total_count_boxes
                existing_none_row.location = locations
                existing_none_row.sum_points = sum_points

            else:
                existing_row = session.query(ReestrTable).filter(ReestrTable.num_th == num_th).first()
                if existing_row:
                    existing_row.total_weight = total_weight
                    existing_row.total_count_boxes = total_count_boxes
                    existing_row.location = locations
                    existing_row.sum_points = sum_points
                else:
                    session.add(ReestrTable(num_th=num_th, total_weight=total_weight, sum_points=sum_points,
                                            total_count_boxes=total_count_boxes, location=locations))
        session.commit()
        print("ReestrTable updated successfully")
        return sorted_total_weights, details_by_num_th

    except Exception as e:
        print("An error occurred:", str(e))


def get_highest_num_th():
    highest_num_th = session.query(ReestrTable).order_by(ReestrTable.num_th.desc()).first()
    if not highest_num_th.num_th:
        highest_num_th = session.query(ReestrTable).order_by(ReestrTable.num_th.desc()).offset(1).limit(1).first()
    print(highest_num_th.num_th)
    return highest_num_th


def get_minimal_num_th():
    # Найти минимальный num_th в таблице ReestrTable
    minimal_num_th = session.query(ReestrTable).order_by(ReestrTable.num_th.asc()).first()
    return minimal_num_th


def increment_num_th(num_th):
    num_parts = num_th.split('-')
    num = int(num_parts[1])
    incremented_num = num + 1
    return f'{num_parts[0]}-{incremented_num:06d}'


def add_num_th_6(result, total_weights_by_num_th, details_by_num_th):
    try:
        if result:
            overweight_in_location_6 = [num_th for num_th, data in total_weights_by_num_th if
                                        '6' in data['locations'] and data['total_weight'] > 3100]
            overweight_code_tt = []
            if overweight_in_location_6:
                for num_th in overweight_in_location_6:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '6':
                                if total_weight > 3100:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_6:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def add_num_th_5(result, total_weights_by_num_th, details_by_num_th):
    try:
        if result:
            overweight_in_location_5 = [num_th for num_th, data in total_weights_by_num_th if
                                        '5' in data['locations'] and data['total_weight'] > 3100]
            overweight_code_tt = []
            if overweight_in_location_5:
                for num_th in overweight_in_location_5:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '4':
                                if total_weight > 3100:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_5:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def check_2_on_12(total_weights_by_num_th, details_by_num_th):
    try:
        location_2_data = [(
            num_th, data['total_weight']) for num_th, data in total_weights_by_num_th if '2' in data['locations']]
        min_num_th = min(location_2_data, key=lambda x: x[1])[0]
        max_num_th = max(location_2_data, key=lambda x: x[1])[0]

        print(min_num_th)
        print(max_num_th)

        code_tt_count_by_num_th = {}

        for num_th, _ in location_2_data:
            code_tt_count_by_num_th[num_th] = 0

        for num_th, details in details_by_num_th.items():
            if num_th in code_tt_count_by_num_th:
                code_tt_count_by_num_th[num_th] += len(details)

        for num_th, count in code_tt_count_by_num_th.items():
            print(f"num_th: {num_th}, code_tt count: {count}")

        overnumber_code_tt = []
        if code_tt_count_by_num_th.get(max_num_th, 0) > 12:
            sorted_items = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'])
            for item in sorted_items:
                if code_tt_count_by_num_th.get(max_num_th, 0) > 12:
                    code_tt_count_by_num_th[max_num_th] -= 1
                    overnumber_code_tt.append(item['code_tt'])
            print(f"Removed code_tt from {max_num_th}: {overnumber_code_tt}")
        if overnumber_code_tt:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overnumber_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
        session.commit()

    except Exception as e:
        print(f"An error occurred: {e}")
        return False


def check_12_th_2():
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        location_2_data = [(num_th, data['total_weight']) for num_th, data in total_weights_by_num_th
                           if '2' in data['locations'] and data['total_weight'] > 1700]
        min_num_th = min(location_2_data, key=lambda x: x[1])[0]
        max_num_th = max(location_2_data, key=lambda x: x[1])[0]

        print(min_num_th)
        print(max_num_th)

        code_tt_count_by_num_th = {}

        for num_th, _ in location_2_data:
            code_tt_count_by_num_th[num_th] = 0

        for num_th, details in details_by_num_th.items():
            if num_th in code_tt_count_by_num_th:
                code_tt_count_by_num_th[num_th] += len(details)

        for num_th, count in code_tt_count_by_num_th.items():
            print(f"num_th: {num_th}, code_tt count: {count}")

        num_th_with_max_count = max(code_tt_count_by_num_th, key=code_tt_count_by_num_th.get)
        max_count = code_tt_count_by_num_th[num_th_with_max_count]

        num_th_with_min_count = min(code_tt_count_by_num_th, key=code_tt_count_by_num_th.get)
        min_count = code_tt_count_by_num_th[num_th_with_max_count]

        print(f"num_th with max code_tt count: {num_th_with_max_count}, code_tt count: {max_count}")

        overnumber_code_tt = []
        if code_tt_count_by_num_th.get(num_th_with_max_count, 0) > 12:
            sorted_items = sorted(details_by_num_th[num_th_with_max_count], key=lambda x: x['weight'])
            for item in sorted_items:
                if code_tt_count_by_num_th.get(num_th_with_max_count, 0) > 11:
                    code_tt_count_by_num_th[num_th_with_max_count] -= 1
                    overnumber_code_tt.append(item['code_tt'])
            print(f"Removed code_tt from {num_th_with_max_count}: {overnumber_code_tt}")
        if overnumber_code_tt:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overnumber_code_tt)).update({'num_th': num_th_with_min_count}, synchronize_session='fetch')
        session.commit()

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        overweight_code_tt = []
        sorted_items = sorted(details_by_num_th[num_th_with_min_count], key=lambda x: x['weight'], reverse=True)

        if num_th_with_min_count:
            sorted_items = sorted(details_by_num_th[num_th_with_min_count], key=lambda x: x['weight'], reverse=True)
            total_weight = next(
                data['total_weight'] for num, data in total_weights_by_num_th if num == num_th_with_min_count)

            for item in sorted_items:
                if item.get('location') == '2' or '1':
                    if total_weight > 3300:
                        total_weight -= item['weight']
                        overweight_code_tt.append(item['code_tt'])
                    else:
                        break

        if overweight_code_tt:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overnumber_code_tt)).update({'num_th': num_th_with_max_count}, synchronize_session='fetch')
        session.commit()

    except Exception as e:
        print(f"An error occurred: {e}")
        return False

#     overweight_code_tt = []
#     if overweight_in_location_2:
#         for num_th in overweight_in_location_2:
#             if num_th in details_by_num_th:
#                 sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
#                 total_weight = next(
#                     data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)
#
#                 for item in sorted_items:
#                     if item.get('location') == '4':
#                         if total_weight > 3100:
#                             total_weight -= item['weight']
#                             overweight_code_tt.append(item['code_tt'])
#                         else:
#                             break
# except Exception as e:
#     print(f"An error occurred: {str(e)}")


def add_num_th_2(result, total_weights_by_num_th, details_by_num_th):
    try:
        if result:
            overweight_in_location_2_1 = [num_th for num_th, data in total_weights_by_num_th if
                                          ('2' in data['locations'] or '1' in data['locations']) and
                                          data['total_weight'] > 3100]

            overweight_code_tt = []
            if overweight_in_location_2_1:
                for num_th in overweight_in_location_2_1:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '1' or item.get('location') == '2':
                                if total_weight > 3100:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_2_1:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def insert_for_7_with_nedoves(total_weights_by_num_th, details_by_num_th):
    print("Start insert_for_7_with_nedoves")
    try:
        overweight_in_location_7 = [num_th for num_th, data in total_weights_by_num_th if
                                    '7' in data['locations'] and data['total_weight'] > 3100]
        only_7_with_nedoves = [num_th for num_th, details in details_by_num_th.items() if
                               all(item['location'] == '7' for item in details) and
                               sum(item['weight'] for item in details) < 2900]
        print("only_7_with_nedoves:", only_7_with_nedoves[0])
        overweight_code_tt = []
        if overweight_in_location_7:
            print("overweight_in_location_7:", overweight_in_location_7)
            for num_th in overweight_in_location_7:
                if only_7_with_nedoves[0]:
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'], reverse=True)
                    try:
                        total_weight = next(
                            data[1]['total_weight'] for data in total_weights_by_num_th if data[0] == only_7_with_nedoves[0])
                    except Exception as e:
                        print(f"An error occurred: {str(e)}")
                        import traceback
                        traceback.print_exc()
                    print("Processing num_th:", num_th)
                    print("Initial total_weight:", total_weight)
                    for item in sorted_items:
                        if item.get('location') == '7':
                            if 2900 <= total_weight <= 3200:
                                break
                            total_weight += item['weight']
                            if 2900 <= total_weight <= 3200:
                                break
                            overweight_code_tt.append(item['code_tt'])
                            print("Added code_tt:", item['code_tt'])
                            print("Updated total_weight:", total_weight)
                            if 2900 <= total_weight <= 3200:
                                break

        print(f"точки для вставки: {overweight_code_tt}")
        if overweight_code_tt and only_7_with_nedoves:
            num_th_to_set = only_7_with_nedoves[0]
            print("num_th_to_set:", num_th_to_set)
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update(
                {'num_th': num_th_to_set}, synchronize_session='fetch')
            session.commit()
        only_7_with_nedoves = only_7_with_nedoves[0]
        total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                           func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
            .filter(AddressTable.num_th == only_7_with_nedoves).group_by(AddressTable.num_th).first()

        if total_weight_query:
            total_weight, total_count_boxes = total_weight_query
            print(f"Total Weight for {only_7_with_nedoves}: {total_weight}, Total Count Boxes: {total_count_boxes}")
            session.query(ReestrTable).filter(ReestrTable.num_th == only_7_with_nedoves).update(
                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
            session.commit()

        for num_th in overweight_in_location_7:
            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()
            print("total_weight_query:", total_weight_query)

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                    synchronize_session='fetch')
                session.commit()

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def add_num_th_7(result):
    try:
        if result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            overweight_in_location_7 = [num_th for num_th, data in total_weights_by_num_th if
                                        '7' in data['locations'] and data['total_weight'] > 3100]

            overweight_code_tt = []
            if overweight_in_location_7:
                for num_th in overweight_in_location_7:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'], reverse=True)
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '6':
                                if total_weight > 3200:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th and highest_num_th.num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            else:
                print(f"We have trable with {highest_num_th}")

            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_7:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_8_0(result, total_weights_by_num_th, details_by_num_th):
    print("Start correct_8_0")
    code_tt_0 = []
    code_tt_8_49 = []
    code_tt_8_4 = []
    code_tt_8_3 = []
    code_tt_8_2 = []
    code_tt_8_1 = []

    try:
        if result == False:
            nedoves_in_location_8_0 = [num_th for num_th, data in total_weights_by_num_th if '8' in data['locations']
                                       and '0' in data['locations'] and data['total_weight'] < 2900]

            code_tt_0.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '0'])

            if code_tt_0 is not None:
                print("code_tt_0:", code_tt_0)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_0)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            code_tt_8_49.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '8' and item.get('priority') == 49])

            if code_tt_8_49 is not None:
                print("code_tt_8_49:", code_tt_8_49)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_8_49)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            code_tt_8_4.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '8' and item.get('priority') == 4])

            if code_tt_8_4 is not None:
                print("code_tt_8_4:", code_tt_8_4)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_8_4)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            code_tt_8_3.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '8' and item.get('priority') == 3])

            if code_tt_8_3 is not None:
                print("code_tt_8_3:", code_tt_8_3)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_8_3)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            code_tt_8_2.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '8' and item.get('priority') == 2])

            if code_tt_8_2 is not None:
                print("code_tt_8_2:", code_tt_8_2)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_8_2)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            code_tt_8_1.extend(
                [item['code_tt'] for item in details_by_num_th[nedoves_in_location_8_0[0]] if
                 item.get('location') == '8' and item.get('priority') == 1])

            if code_tt_8_1 is not None:
                print("code_tt_8_1:", code_tt_8_1)
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(code_tt_8_1)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            if nedoves_in_location_8_0:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th.in_(nedoves_in_location_8_0)).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(
                        f"Total Weight for {nedoves_in_location_8_0}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                    session.query(ReestrTable).filter(ReestrTable.num_th.in_(nedoves_in_location_8_0)).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')
                    session.commit()
                else:
                    for num_th in nedoves_in_location_8_0:
                        session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                            {'total_weight': 0, 'total_count_boxes': 0}, synchronize_session='fetch')
                    session.commit()
                    print(f"у {nedoves_in_location_8_0} выставляем 0 в ReestreTable")

        #return code_tt_8_49, code_tt_8_4, code_tt_8_3, code_tt_8_2, code_tt_8_1
        print("End correct_8_0")

    except Exception as e:
        print("Error in correct_8:", e)


def correct_8_1(total_weights_by_num_th, details_by_num_th):
    print("Start correct_8_1")
    max_num_th_8 = None
    max_num_th_9 = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th_8 = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_1 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 1]
        code_tt_8_1 = [item['code_tt'] for item in filtered_8_1]
        weight_code_tt_8_1 = [item['weight'] for item in filtered_8_1]
        print(weight_code_tt_8_1)

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
                                 if '9' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_9:", nedoves_in_location_9)
        if nedoves_in_location_9:
            max_num_th_9 = max(
                nedoves_in_location_9,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        if max_num_th_9:
            max_num_th = max_num_th_9
        else:
            max_num_th = max_num_th_8

        if max_num_th:
            if code_tt_8_1:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_1)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_1)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

        print("End correct_8_2")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_8_2(total_weights_by_num_th, details_by_num_th):
    print("Start correct_8_2")
    max_num_th_8 = None
    max_num_th_9 = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th_8 = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_2 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 2]
        code_tt_8_2 = [item['code_tt'] for item in filtered_8_2]
        weight_code_tt_8_2 = [item['weight'] for item in filtered_8_2]
        print(weight_code_tt_8_2)

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
                                 if '9' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_9:", nedoves_in_location_9)
        if nedoves_in_location_9:
            max_num_th_9 = max(
                nedoves_in_location_9,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        if max_num_th_9:
            max_num_th = max_num_th_9
        else:
            max_num_th = max_num_th_8

        if max_num_th:
            if code_tt_8_2:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_2)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_2)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

        print("End correct_8_2")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_4_to_num_th_8(result):
    print("Start moving_6_4_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_4 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 4]
            code_tt_6_4 = [item['code_tt'] for item in filtered_6_4]
            weight_code_tt_6_4 = [item['weight'] for item in filtered_6_4]

            if max_num_th_8:
                if code_tt_6_4:
                    print(weight_code_tt_6_4)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_4)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_4)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes'))\
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End moving_6_4_to_num_th_8")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_49_to_num_th_8(result):
    print("Start moving_6_49_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_49 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 49]
            code_tt_6_49 = [item['code_tt'] for item in filtered_6_49]
            weight_code_tt_6_49 = [item['weight'] for item in filtered_6_49]

            if max_num_th_8:
                if code_tt_6_49:
                    print(weight_code_tt_6_49)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_49)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_49)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_49")

    except Exception as e:
        print(f"An error occurred: {str(e)}")



def moving_6_3_to_num_th_8(result):
    print("Start moving_6_3_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_3 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 3]
            code_tt_6_3 = [item['code_tt'] for item in filtered_6_3]
            weight_code_tt_6_3 = [item['weight'] for item in filtered_6_3]

            if max_num_th_8:
                if code_tt_6_3:
                    print(weight_code_tt_6_3)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_3)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_3)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_3")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_2_to_num_th_8(result):
    print("Start moving_6_2_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 3000 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_2 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 2]
            code_tt_6_2 = [item['code_tt'] for item in filtered_6_2]
            weight_code_tt_6_2 = [item['weight'] for item in filtered_6_2]

            if max_num_th_8:
                if code_tt_6_2:
                    print(weight_code_tt_6_2)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_2)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_2)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_2")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_1_to_num_th_8(result):
    print("Start moving_6_1_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 3000 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_1 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 1]
            code_tt_6_1 = [item['code_tt'] for item in filtered_6_1]
            weight_code_tt_6_1 = [item['weight'] for item in filtered_6_1]

            if max_num_th_8:
                if code_tt_6_1:
                    print(weight_code_tt_6_1)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_1)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_1)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_1")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_44_to_num_th_8(result):
    print("Start moving_6_44_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 3300 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_44 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 44]
            code_tt_6_44 = [item['code_tt'] for item in filtered_6_44]
            weight_code_tt_6_44 = [item['weight'] for item in filtered_6_44]

            if max_num_th_8:
                if code_tt_6_44:
                    print(weight_code_tt_6_44)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_44)
                    if total_weight_with_new_items <= 3300:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_44)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_44")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_6_to_num_th_8(result):
    print("Start moving_6_6_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 3000 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_6 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 6]
            code_tt_6_6 = [item['code_tt'] for item in filtered_6_6]
            weight_code_tt_6_6 = [item['weight'] for item in filtered_6_6]

            if max_num_th_8:
                if code_tt_6_6:
                    print(weight_code_tt_6_6)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_6)
                    if total_weight_with_new_items <= 3200:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_6)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_6")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def moving_6_45_to_num_th_8(result):
    print("Start moving_6_45_to_num_th_8")
    max_num_th_8 = None
    try:
        if not result:
            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                     if '8' in data['locations'] and data['total_weight'] < 3000 and num_th is not None]

            if nedoves_in_location_8:
                print("nedoves_in_location_8:", nedoves_in_location_8)
                max_num_th_8 = max(nedoves_in_location_8, key=lambda num_th: next(data['total_weight'] for num,
                data in total_weights_by_num_th if num == num_th), default=float('inf'))

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

            filtered_6_45 = [item for item in details_by_num_th.get(None, []) if item.get(
                'location') == '6' and item.get('priority') == 45]
            code_tt_6_45 = [item['code_tt'] for item in filtered_6_45]
            weight_code_tt_6_45 = [item['weight'] for item in filtered_6_45]

            if max_num_th_8:
                if code_tt_6_45:
                    print(weight_code_tt_6_45)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
                    total_weight_with_new_items = total_weight + sum(weight_code_tt_6_45)
                    if total_weight_with_new_items <= 3200:
                        session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_45)).update(
                            {'num_th': max_num_th_8}, synchronize_session='fetch')
                        session.commit()

                        total_weight_query = session.query(
                            func.sum(AddressTable.weight).label('total_weight'),
                            func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th_8}: {total_weight}, Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End correct_6_45")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_0_3_old(total_weights_by_num_th, details_by_num_th):
    print("Start correct_0_3")
    max_num_th_8 = None
    max_num_th_9 = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th_8 = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_0_3 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '0' and item.get('priority') == 3]
        code_tt_0_3 = [item['code_tt'] for item in filtered_0_3]
        weight_code_tt_0_3 = [item['weight'] for item in filtered_0_3]
        print(weight_code_tt_0_3)

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
                                 if '9' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_9:", nedoves_in_location_9)
        if nedoves_in_location_9:
            max_num_th_9 = max(
                nedoves_in_location_9,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        if max_num_th_9:
            max_num_th = max_num_th_9
        else:
            max_num_th = max_num_th_8

        if max_num_th:
            if code_tt_0_3:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_0_3)
                if total_weight_with_new_items <= 3050:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_0_3)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

        print("End correct_0_3")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_0_2(total_weights_by_num_th, details_by_num_th):
    print("Start correct_0_2")
    min_num_th_2 = None
    try:
        nedoves_in_location_2 = [num_th for num_th, data in total_weights_by_num_th
                                 if '2' in data['locations'] and data['total_weight'] < 2700 and num_th is not None]

        print("nedoves_in_location_2:", nedoves_in_location_2)
        if nedoves_in_location_2:
            min_num_th_2 = min(
                nedoves_in_location_2,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_0_2 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '0' and item.get('priority') == 2]
        code_tt_0_2 = [item['code_tt'] for item in filtered_0_2]
        weight_code_tt_0_2 = [item['weight'] for item in filtered_0_2]
        print(weight_code_tt_0_2)

        if min_num_th_2:
            if code_tt_0_2:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == min_num_th_2)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_0_2)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_0_2)).update(
                        {'num_th': min_num_th_2}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == min_num_th_2).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {min_num_th_2}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == min_num_th_2).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

        print("End correct_0_2")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_0_1(total_weights_by_num_th, details_by_num_th):
    print("Start correct_0_1")
    max_num_th_8 = None
    max_num_th_9 = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th_8 = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_0_1 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '0' and item.get('priority') == 1]
        code_tt_0_1 = [item['code_tt'] for item in filtered_0_1]
        weight_code_tt_0_1 = [item['weight'] for item in filtered_0_1]
        print(weight_code_tt_0_1)

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
                                 if '9' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_9:", nedoves_in_location_9)
        if nedoves_in_location_9:
            max_num_th_9 = max(
                nedoves_in_location_9,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        if max_num_th_9:
            max_num_th = max_num_th_9
        else:
            max_num_th = max_num_th_8

        if max_num_th:
            if code_tt_0_1:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_0_1)
                if total_weight_with_new_items <= 3050:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_0_1)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

        print("End correct_0_1")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


# def correct_8_2(total_weights_by_num_th, details_by_num_th):
#     print("Start correct_8")
#     max_num_th_8 = None
#     max_num_th_9 = None
#     try:
#         nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
#                                  if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]
#
#         print("nedoves_in_location_8:", nedoves_in_location_8)
#         if nedoves_in_location_8:
#             max_num_th_8 = max(
#                 nedoves_in_location_8,
#                 key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
#                 default=float('inf'))
#
#         # Получаем code_tt, у которых num_th равен None
#         code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
#         filtered_8_2 = [item for item in details_by_num_th.get(None, []) if item.get(
#             'location') == '8' and item.get('priority') == 2]
#         code_tt_8_2 = [item['code_tt'] for item in filtered_8_2]
#         weight_code_tt_8_2 = [item['weight'] for item in filtered_8_2]
#         print(weight_code_tt_8_2)
#
#         nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
#                                  if '9' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]
#
#         print("nedoves_in_location_9:", nedoves_in_location_8)
#         elif:
#          if nedoves_in_location_9:
#             max_num_th_9 = max(
#                 nedoves_in_location_9,
#                 key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
#                 default=float('inf'))
#
#         # Получаем code_tt, у которых num_th равен None
#         code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
#         filtered_8_2 = [item for item in details_by_num_th.get(None, []) if item.get(
#             'location') == '8' and item.get('priority') == 2]
#         code_tt_8_2 = [item['code_tt'] for item in filtered_8_2]
#         weight_code_tt_8_2 = [item['weight'] for item in filtered_8_2]
#         print(weight_code_tt_8_2)
#         if max_num_th_9:
#             if code_tt_8_2:
#                 total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
#                 total_weight_with_new_items = total_weight + sum(weight_code_tt_8_2)
#                 if total_weight_with_new_items <= 3200:
#                     session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_2)).update(
#                         {'num_th': max_num_th_8}, synchronize_session='fetch')
#                     session.commit()
#
#                     total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
#                                                        func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
#                         .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()
#
#                     if total_weight_query:
#                         total_weight, total_count_boxes = total_weight_query
#                         print(f"Total Weight for {max_num_th_8}: {total_weight}, Total Count Boxes: {total_count_boxes}")
#                         session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
#                             {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
#                             synchronize_session='fetch')
#                         session.commit()
#         elif:
#             if max_num_th_9:
#                 if code_tt_8_2:
#                     total_weight = next(
#                         data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th_8)
#                     total_weight_with_new_items = total_weight + sum(weight_code_tt_8_2)
#                     if total_weight_with_new_items <= 3200:
#                         session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_2)).update(
#                             {'num_th': max_num_th_8}, synchronize_session='fetch')
#                         session.commit()
#
#                         total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
#                                                            func.sum(AddressTable.count_boxes).label(
#                                                                'total_count_boxes')) \
#                             .filter(AddressTable.num_th == max_num_th_8).group_by(AddressTable.num_th).first()
#
#                         if total_weight_query:
#                             total_weight, total_count_boxes = total_weight_query
#                             print(
#                                 f"Total Weight for {max_num_th_8}: {total_weight}, Total Count Boxes: {total_count_boxes}")
#                             session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th_8).update(
#                                 {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
#                                 synchronize_session='fetch')
#                             session.commit()
#             print("End correct_8_2")
#
#     except Exception as e:
#         print(f"An error occurred: {str(e)}")


def correct_8_3(total_weights_by_num_th, details_by_num_th):
    print("Start correct_8")
    max_num_th = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_3 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 3]
        code_tt_8_3 = [item['code_tt'] for item in filtered_8_3]
        weight_code_tt_8_3 = [item['weight'] for item in filtered_8_3]
        print(weight_code_tt_8_3)

        if max_num_th:
            if code_tt_8_3:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_3)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_3)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()
            print("End correct_8_3")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_8_4(total_weights_by_num_th, details_by_num_th):
    print("Start correct_8")
    max_num_th = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_4 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 4]
        code_tt_8_4 = [item['code_tt'] for item in filtered_8_4]
        weight_code_tt_8_4 = [item['weight'] for item in filtered_8_4]
        print(weight_code_tt_8_4)

        if max_num_th:
            if code_tt_8_4:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_4)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_4)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                        session.commit()
            print("End correct_8_4")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_0_1(total_weights_by_num_th, details_by_num_th):
    print("Start correct_0_1")
    max_num_th = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_49 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 49]
        code_tt_8_49 = [item['code_tt'] for item in filtered_8_49]
        weight_code_tt_8_49 = [item['weight'] for item in filtered_8_49]
        print(weight_code_tt_8_49)
        if max_num_th:
            if code_tt_8_49:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_49)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_49)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                        session.commit()

            # if code_tt_8_49:
            #     total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
            #     total_weight_with_new_items = total_weight + sum(weight_code_tt_8_49)
            #     if total_weight_with_new_items <= 3200:
            #         session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_49)).update(
            #             {'num_th': max_num_th}, synchronize_session='fetch')
            #
            #         session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
            #             {'total_weight': func.sum(AddressTable.weight),
            #              'total_count_boxes': func.sum(AddressTable.count_boxes)},
            #             synchronize_session='fetch')
            #
            #         session.commit()
            #
            #         print(f"Total Weight for {max_num_th}: {total_weight_with_new_items}")

            print("End correct_8_49")

    except Exception as e:
        print(f"An error occurred: {str(e)}")



def correct_8_49(total_weights_by_num_th, details_by_num_th):
    print("Start correct_8_49")
    max_num_th = None
    try:
        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th
                                 if '8' in data['locations'] and data['total_weight'] < 2900 and num_th is not None]

        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            max_num_th = max(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        # Получаем code_tt, у которых num_th равен None
        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
        filtered_8_49 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '8' and item.get('priority') == 49]
        code_tt_8_49 = [item['code_tt'] for item in filtered_8_49]
        weight_code_tt_8_49 = [item['weight'] for item in filtered_8_49]
        print(weight_code_tt_8_49)
        if max_num_th:
            if code_tt_8_49:
                total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_8_49)
                if total_weight_with_new_items <= 3200:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_49)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                        session.commit()

            # if code_tt_8_49:
            #     total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
            #     total_weight_with_new_items = total_weight + sum(weight_code_tt_8_49)
            #     if total_weight_with_new_items <= 3200:
            #         session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_49)).update(
            #             {'num_th': max_num_th}, synchronize_session='fetch')
            #
            #         session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
            #             {'total_weight': func.sum(AddressTable.weight),
            #              'total_count_boxes': func.sum(AddressTable.count_boxes)},
            #             synchronize_session='fetch')
            #
            #         session.commit()
            #
            #         print(f"Total Weight for {max_num_th}: {total_weight_with_new_items}")

            print("End correct_8_49")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_max_to_min(actually_info):
    print("Start correct_max_to_min")
    print(actually_info)
    max_num_th = None
    null_num_th = None
    try:
        null_num_th = [num_th for num_th, weight in actually_info.items() if weight == 0]

        max_num_th = max(actually_info.keys(), key=lambda x: int(x.split('-')[1])) if actually_info else None

        print("num_th with total_weight = 0:", null_num_th)
        print("num_th with the highest numerical value:", max_num_th)

        if null_num_th and max_num_th and null_num_th[0] < max_num_th:

            session.query(ReestrTable).filter(ReestrTable.num_th == null_num_th[0]).delete(synchronize_session='fetch')
            session.commit()

            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                {'num_th': null_num_th[0]},
                synchronize_session='fetch')
            session.commit()

            session.query(AddressTable).filter(AddressTable.num_th == max_num_th).update(
                {'num_th': null_num_th[0]},
                synchronize_session='fetch')
            session.commit()

            print(f"Updated ReestrTable.num_th from {max_num_th} to {null_num_th[0]}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def correct_1(total_weights_by_num_th, details_by_num_th):
    print("Start correct_1")
    min_num_th = None
    try:
        norm_in_location_1 = [num_th for num_th, data in total_weights_by_num_th if
                                    '1' in data['locations'] and data['total_weight'] > 2900]

        nedoves_in_location_1 = [num_th for num_th, data in total_weights_by_num_th if
                                    '1' in data['locations'] and data['total_weight'] < 2700]

        print("norm_in_location_1:", norm_in_location_1)
        print("nedoves_in_location_1:", nedoves_in_location_1)

        # if norm_in_location_1:
        #     min_num_th = max(
        #         norm_in_location_1,
        #         key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
        #         default=float('inf'))
        #     print("Num_th with max total_weight in nedoves_in_location_1:", min_num_th)

        code_tt_for_nedoves = []
        if norm_in_location_1:
            for num_th in norm_in_location_1:
                if num_th in details_by_num_th:
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                    if sorted_items:  # Ensure there are items to process
                        for item in sorted_items:
                            if item['weight'] >= 200.0:
                                code_tt_for_nedoves.append(item['code_tt'])
                                break

        print("code_tt_for_nedoves:", code_tt_for_nedoves)

        if nedoves_in_location_1 is not None:
            for num_th in nedoves_in_location_1:
                session.query(AddressTable).filter(AddressTable.code_tt.in_(
                    code_tt_for_nedoves)).update({'num_th': num_th}, synchronize_session='fetch')
                session.commit()

                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {min_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                    session.commit()

        if norm_in_location_1 is not None:
            print("norm_in_location_1:", norm_in_location_1)
            for num_th in norm_in_location_1:
                print("norm_in_location_1:", norm_in_location_1)
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()
            print("End correct_1")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def share_weight_8(total_weights_by_num_th, details_by_num_th):
    print("Start share_weight_8")
    min_num_th = None
    try:
        overweight_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                    '8' in data['locations'] and data['total_weight'] > 3100]

        nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                    '8' in data['locations'] and data['total_weight'] < 2900]

        print("overweight_in_location_8:", overweight_in_location_8)
        print("nedoves_in_location_8:", nedoves_in_location_8)
        if nedoves_in_location_8:
            min_num_th = min(
                nedoves_in_location_8,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Num_th with minimum total_weight in nedoves_in_location_8:", min_num_th)

        overweight_code_tt = []
        if overweight_in_location_8:
            for num_th in overweight_in_location_8:
                if num_th in details_by_num_th:
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                    for item in sorted_items:
                        if item.get('location') == '8':
                            if total_weight > 3050:
                                total_weight -= item['weight']
                                overweight_code_tt.append(item['code_tt'])
                            else:
                                break

        print("overweight_code_tt:", overweight_code_tt)
        if min_num_th is not None:
            print(min_num_th)
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == min_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {min_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == min_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_8:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()
            print("End share_weight_8")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def share_weight_9(total_weights_by_num_th, details_by_num_th):
    print("Start share_weight_9")
    min_num_th = None
    try:

        check_12 = {}

        for num_th, data in details_by_num_th.items():
            count_code_tt = len(data)
            if count_code_tt >= 12:
                check_12[num_th] = count_code_tt

        print("check_12:", check_12)

        overweight_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                    '9' in data['locations'] and data['total_weight'] > 3100]

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                    '9' in data['locations'] and data['total_weight'] < 3100]

        nedoves_in_location_9_and_check_12 = [num_th for num_th in nedoves_in_location_9 if num_th not in check_12]

        print("nedoves_in_location_9_and_check_12:", nedoves_in_location_9_and_check_12)
        print("overweight_in_location_9:", overweight_in_location_9)
        print("nedoves_in_location_9:", nedoves_in_location_9)
        if nedoves_in_location_9_and_check_12:
            min_num_th = min(
                nedoves_in_location_9_and_check_12,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Num_th with minimum total_weight in nedoves_in_location_9:", min_num_th)

        overweight_code_tt = []
        if overweight_in_location_9:
            for num_th in overweight_in_location_9:
                if num_th in details_by_num_th:
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'], reverse=True)
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                    for item in sorted_items:
                        if item.get('location') == '9':
                            if total_weight > 3100:
                                total_weight -= item['weight']
                                overweight_code_tt.append(item['code_tt'])
                            else:
                                break

        print("overweight_code_tt:", overweight_code_tt)
        if min_num_th is not None:
            print(min_num_th)
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == min_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {min_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == min_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_9:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()
            print("End share_weight_9")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_9(total_weights_by_num_th, details_by_num_th):
    print("Start refactor_9")
    try:
        overweight_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                    '9' in data['locations'] and data['total_weight'] > 3100]

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                    '9' in data['locations'] and data['total_weight'] < 2700]
        print(nedoves_in_location_9)

        location_9_6 = [num_th for num_th, data in total_weights_by_num_th if
                            '9' in data['locations'] and '6' in data['locations']]

        overweight_code_tt = []

        if nedoves_in_location_9:
            for num_th in overweight_in_location_9:
                if num_th in details_by_num_th:
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                    total_weight = next(
                        data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)
                    print(total_weight)
                    for item in sorted_items:
                        if total_weight >= 3300:
                            if item['weight'] > 400:
                                total_weight -= item['weight']
                                overweight_code_tt.append(item['code_tt'])
                        else:
                            break
        if overweight_code_tt:
            for num_th in nedoves_in_location_9:
                print(num_th)
                session.query(AddressTable).filter(AddressTable.code_tt.in_(
                    overweight_code_tt)).update({'num_th': num_th}, synchronize_session='fetch')
                session.commit()


        # if location_6_code_tt:
        #     session.query(AddressTable).filter(
        #         AddressTable.code_tt.in_(location_6_code_tt)).update({'num_th': None}, synchronize_session='fetch')
        #     session.commit()

        print("End refactor_9_6")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_9_nedoves_3_12():
    print("Start refactor_9_nedoves_3_12")
    min_num_th = None
    max_num_th = None
    total_weight_nedoves_norm = None
    try:
        # total_weights_by_num_th = dict(total_weights_by_num_th)
        # nedoves_location_6 = [num_th for num_th, data in total_weights_by_num_th.items() if
        #                       ('6' in data['locations'] and ('7' in data['locations'] or '8' in data['locations']))
        #                       and data['total_weight'] < 3100]

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        location_9_6 = [num_th for num_th, data in total_weights_by_num_th if
                            '9' in data['locations'] and '6' in data['locations']]

        nedoves_9_with_12 = [num_th for num_th, data in total_weights_by_num_th if
                            '9' in data['locations'] and data['sum_points'] >= 12 and data['total_weight'] < 2800]

        nedoves_9_norm = [num_th for num_th, data in total_weights_by_num_th if
                          '9' in data['locations'] and data['total_weight'] < 2800 and num_th not in nedoves_9_with_12]

        if nedoves_9_with_12:
            max_num_th = max(nedoves_9_with_12,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("nedoves_9_with_12:", max_num_th)

        if nedoves_9_norm:
            min_num_th = min(nedoves_9_norm,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("nedoves_9_norm:", max_num_th)
            if min_num_th in details_by_num_th:
                total_weight_min_th = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == min_num_th)

        code_tt_9 = []
        if max_num_th and min_num_th:
            if max_num_th in details_by_num_th:
                sorted_items = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'])
                total_weight = total_weight_min_th
                for item in sorted_items:
                    if item.get('location') == '9':
                        if total_weight < 2800:
                            total_weight += item['weight']
                            code_tt_9.append(item['code_tt'])
                        else:
                            break

        if nedoves_9_norm:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                code_tt_9)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()
            logging.debug("num_th updated successfully for location 9 code_tt.")

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]

        filtered_6_2 = [item for item in details_by_num_th.get(None, []) if item.get(
            'location') == '6' and item.get('priority') == 2]
        code_tt_6_2 = [item['code_tt'] for item in filtered_6_2]
        weight_code_tt_6_2 = [item['weight'] for item in filtered_6_2]

        print("End refactor_9_nedoves_3")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_9_nedoves_3_overweight():
    print("Start refactor_9_nedoves")
    min_num_th = None
    max_num_th = None
    sum_address_min_num_th = None

    try:
        # total_weights_by_num_th = dict(total_weights_by_num_th)
        # nedoves_location_6 = [num_th for num_th, data in total_weights_by_num_th.items() if
        #                       ('6' in data['locations'] and ('7' in data['locations'] or '8' in data['locations']))
        #                       and data['total_weight'] < 3100]

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        overweight_3400 = [num_th for num_th, data in total_weights_by_num_th if
                                 '9' in data['locations'] and 3500 >= data['total_weight'] >= 3400]

        if overweight_3400:
            max_num_th = max(overweight_3400,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        with_12_2300 = [num_th for num_th, data in total_weights_by_num_th if
                            '9' in data['locations'] and data['sum_points'] >= 12 and data['total_weight'] < 2300]

        if with_12_2300:
            min_num_th = min(with_12_2300,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))

        overwight_9_code_tt = []
        if with_12_2300 and overweight_3400:
            if max_num_th in details_by_num_th:
                items_in_max_num_th = [item for item in details_by_num_th[max_num_th]]
                sorted_items_max = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'], reverse=True)
                max_weight = sorted_items_max[0]['weight']
                sorted_items_min = sorted(details_by_num_th[min_num_th], key=lambda x: x['weight'])
                min_weight = sorted_items_min[0]['weight']
                total_weight = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)

                print(sorted_items_max[0]['weight'])
                print(sorted_items_max[1]['weight'])
                print(sorted_items_max[2]['weight'])

                overwight_9_code_tt.append(sorted_items_max[0]['code_tt'])

        if overwight_9_code_tt:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overwight_9_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()
            logging.debug("num_th updated successfully for location 6 code_tt.")

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        print("End refactor_9_nedoves_3")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_9_nedoves_give_6():
    print("Start refactor_9_6")
    min_num_th = None
    max_num_th = None
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()
        total_weights_by_num_th = dict(total_weights_by_num_th)

        nedoves_location_6 = [num_th for num_th, data in total_weights_by_num_th.items() if
                              ('6' in data['locations'] and ('7' in data['locations'] or '8' in data['locations']))
                              and data['total_weight'] < 3100]

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        location_9_6 = [num_th for num_th, data in total_weights_by_num_th if
                            '9' in data['locations'] and '6' in data['locations']]

        if nedoves_location_6:
            min_num_th = min(
                nedoves_location_6,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Min weight num_th location 6:", min_num_th)

        if location_9_6:
            max_num_th = max(
                location_9_6,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Max weight num_th location 9 and 6:", max_num_th)

        location_6_code_tt = []
        if location_9_6:
            if max_num_th in details_by_num_th:
                sorted_items = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'])
                total_weight = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)

                for item in sorted_items:
                    if item.get('location') == '6':
                        if total_weight != 0:
                            total_weight -= item['weight']
                            location_6_code_tt.append(item['code_tt'])
                        else:
                            break

        if location_6_code_tt:
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                location_6_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()
            logging.debug("num_th updated successfully for location 6 code_tt.")

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        print("End refactor_9_6")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_9_6_nedoves(result):
    print("Start refactor_9_6")
    min_num_th = None
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        if not result:
            overweight_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                        '9' in data['locations'] and data['total_weight'] > 3100]

            nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                        '9' in data['locations'] and data['total_weight'] < 2700]

            location_9_6 = [num_th for num_th, data in total_weights_by_num_th if
                                '9' in data['locations'] and '6' in data['locations']]

            if nedoves_in_location_9:
                min_num_th = min(
                    nedoves_in_location_9,
                    key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                    default=float('inf'))
                print("Num_th with maximum total_weight in nedoves_in_location_9:", min_num_th)

            overweight_code_tt_9 = []

            if overweight_in_location_9 and min_num_th:
                for num_th in overweight_in_location_9:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '9':
                                if total_weight > 3004:
                                    total_weight -= item['weight']
                                    overweight_code_tt_9.append(item['code_tt'])
                                else:
                                    break

            if overweight_code_tt_9:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(overweight_code_tt_9)).update({'num_th': min_num_th}, synchronize_session='fetch')
                session.commit()

        print("End refactor_9_6")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def add_num_th_9(result, total_weights_by_num_th, details_by_num_th):
    print("Start add_num_th_9")
    try:
        if result:
            overweight_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                        '9' in data['locations'] and data['total_weight'] > 3100]
            overweight_code_tt = []
            if overweight_in_location_9:
                for num_th in overweight_in_location_9:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '9':
                                if total_weight > 3100:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_9:
                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()
        print("End add_num_th_9")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def add_num_th_8(result, total_weights_by_num_th, details_by_num_th):
    print("Start add_num_th_8")
    try:
        if result:
            overweight_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                        '8' in data['locations'] and data['total_weight'] > 3100]
            overweight_code_tt = []
            if overweight_in_location_8:
                for num_th in overweight_in_location_8:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '8':
                                if total_weight > 3100:
                                    total_weight -= item['weight']
                                    overweight_code_tt.append(item['code_tt'])
                                else:
                                    break

            highest_num_th = get_highest_num_th()
            new_id = str(uuid.uuid4())
            if highest_num_th:
                new_num_th = increment_num_th(highest_num_th.num_th)
                date_th = highest_num_th.date_th
            new_record = ReestrTable(num_th=new_num_th, total_weight=Decimal('0.0'), date_th=date_th, id=new_id)
            session.add(new_record)
            session.commit()
            session.query(AddressTable).filter(AddressTable.code_tt.in_(
                overweight_code_tt)).update({'num_th': new_num_th}, synchronize_session='fetch')
            session.commit()

            total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                               func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                .filter(AddressTable.num_th == new_num_th).group_by(AddressTable.num_th).first()

            if total_weight_query:
                total_weight, total_count_boxes = total_weight_query
                print(f"Total Weight for {new_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                session.query(ReestrTable).filter(ReestrTable.num_th == new_num_th).update(
                    {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
                session.commit()

            for num_th in overweight_in_location_8:

                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                    session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')

            session.commit()
        else:
            print("End add_num_th_8")

        print("End add_num_th_8")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


# def check_2_weight_num_th_9(total_weights_by_num_th, details_by_num_th):
#     print("Start check_2_weight_num_th_9")
#     nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th
#                              if '9' in data['locations'] and data['total_weight'] < 2700 and num_th is not None]
#     code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
#     filtered_9 = [item for item in details_by_num_th.get(None, []) if item.get(
#         'location') == '8' and item.get('priority') < 49]
#     code_tt_9 = [item['code_tt'] for item in filtered_9]
#     weight_code_tt_9 = [item['weight'] for item in filtered_9]
#     if nedoves_in_location_9:
#         if filtered_9:
#             total_weight = next(
#                 data['total_weight'] for num, data in total_weights_by_num_th if num == nedoves_in_location_9)
#             total_weight_with_new_items = total_weight + sum(weight_code_tt_8_49)
#             if total_weight_with_new_items <= 3200:
#                 session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_8_49)).update(
#                     {'num_th': max_num_th}, synchronize_session='fetch')
#                 session.commit()
#
#                 total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
#                                                    func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
#                     .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()
#
#                 if total_weight_query:
#                     total_weight, total_count_boxes = total_weight_query
#                     print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
#                     session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
#                         {'total_weight': total_weight, 'total_count_boxes': total_count_boxes}, synchronize_session='fetch')
#                     session.commit()


def check_weight_num_th_9():
    print("Start check_weight_num_th_9")
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        all_location_9 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations']]

        total_weight_location_9 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_9)

        num_th_in_location_9 = len(all_location_9)

        if num_th_in_location_9 > 0:
            average_weight_location_9 = total_weight_location_9 / num_th_in_location_9

        else:
            average_weight_location_9 = 0.0

            print(f"ТН с location 7 - {all_location_9}. Их текущий суммарный вес - {total_weight_location_9}."
                  f" Их перевес - {average_weight_location_9}")

        if average_weight_location_9 >= 3300:
            return "Overweight", num_th_in_location_9
        elif num_th_in_location_9 > 1 and average_weight_location_9 < 2800:
            return "Underweight", num_th_in_location_9
        else:
            return "Norm", num_th_in_location_9

    except ZeroDivisionError:
        return "Error: Division by zero occurred"


def check_nedoves_num_th_8(total_weights_by_num_th):
    print("Start check_weight_num_th_8")
    try:
        all_location_8 = [num_th for num_th, data in total_weights_by_num_th if '8' in data['locations']]

        total_weight_location_8 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_8)

        num_th_in_location_8 = len(all_location_8)

        if num_th_in_location_8 > 0:
            average_weight_location_8 = total_weight_location_8 - (3100 * num_th_in_location_8)
        else:
            average_weight_location_8 = 0.0

        print("Перевес на location 8:", average_weight_location_8)
        print("End check_weight_num_th_8")

        return 500 <= average_weight_location_8 <= 4000

    except ZeroDivisionError:
        return 0.0


def refactor_8(result):
    print("Start refactor_nedoves_num_th_8")
    max_num_th = None
    try:
        if not result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            min_nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                     '8' in data['locations'] and data['total_weight'] > 1500]

            max_nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                     '8' in data['locations'] and data['total_weight'] < 1500]

            if min_nedoves_in_location_8:
                max_num_th = max(
                    min_nedoves_in_location_8,
                    key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                    default=float('inf'))
                print("Num_th with maximum total_weight in nedoves_in_location_8:", max_num_th)

            location_8_code_tt = []

            if max_nedoves_in_location_8:
                for num_th in max_nedoves_in_location_8:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '8':
                                if total_weight != 0:
                                    total_weight -= item['weight']
                                    location_8_code_tt.append(item['code_tt'])
                                else:
                                    break

            if max_num_th is not None:
                print(max_num_th)
                session.query(AddressTable).filter(AddressTable.code_tt.in_(
                    location_8_code_tt)).update({'num_th': max_num_th}, synchronize_session='fetch')
                session.commit()

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            print("End refactor_nedoves_num_th_8")
            return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def check_nedoves_6_6(result):
    print("Start check_nedoves_6")
    max_num_th = None
    try:
        if not result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_6 = [num_th for num_th, data in total_weights_by_num_th if
                                     '6' in data['locations'] and data['total_weight'] < 2800]

            print(nedoves_in_location_6)

            total_weight_location_6 = sum(
                data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in nedoves_in_location_6)

            total_with_nedoves_6 = len(nedoves_in_location_6)

            if total_with_nedoves_6 == 2:
                for num_th in nedoves_in_location_6:
                    if num_th is not None:
                        max_num_th = num_th
                        print("Num_th is not None in location 6:", max_num_th)

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
            get_6_with_None = [item for item in details_by_num_th.get(None, []) if item.get('location') == '6']
            code_tt_6_None = [item['code_tt'] for item in get_6_with_None]
            weight_code_tt_6_None = [item['weight'] for item in get_6_with_None]
            print(weight_code_tt_6_None)

            if 3400 > total_weight_location_6:
                for num_th in nedoves_in_location_6:
                    if num_th is None:
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                        total_weight_with_new_items = total_weight + sum(weight_code_tt_6_None)
                        if total_weight_with_new_items <= 3200:
                            session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_6_None)).update(
                                {'num_th': max_num_th}, synchronize_session='fetch')
                            session.commit()
                        total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                           func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                            .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                        if total_weight_query:
                            total_weight, total_count_boxes = total_weight_query
                            print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                            session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                                {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                                synchronize_session='fetch')
                            session.commit()

            print("End refactor_nedoves_num_th_8")
            return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def check_nedoves_6_0(result):
    print("Start check_nedoves_6_0")
    max_num_th = None
    try:
        if not result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_6 = [num_th for num_th, data in total_weights_by_num_th if
                                     '6' in data['locations'] and data['total_weight'] < 2800]

            print(nedoves_in_location_6)

            total_weight_location_6 = sum(
                data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in nedoves_in_location_6)

            total_with_nedoves_6 = len(nedoves_in_location_6)

            for num_th in nedoves_in_location_6:
                if num_th is not None:
                    max_num_th = num_th
                    print("Num_th is not None in location 6:", max_num_th)

            # Получаем code_tt, у которых num_th равен None
            code_tt_none = [item['code_tt'] for item in details_by_num_th.get(None, [])]
            get_0_with_None = [item for item in details_by_num_th.get(None, []) if item.get('location') == '0']
            code_tt_0_None = [item['code_tt'] for item in get_0_with_None]
            weight_code_tt_0_None = [item['weight'] for item in get_0_with_None]
            print(weight_code_tt_0_None)

            if max_num_th:
                total_weight = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == max_num_th)
                total_weight_with_new_items = total_weight + sum(weight_code_tt_0_None)
                if total_weight_with_new_items < 3350:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(code_tt_0_None)).update(
                        {'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label(
                                                       'total_count_boxes')) \
                    .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(
                        f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                    session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')
                    session.commit()

            print("End refactor_nedoves_num_th_8")
            return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_8_nedoves(result):
    print("Start refactor_8_nedoves")
    min_num_th = None
    try:
        if not result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            pereves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                        '8' in data['locations'] and data['total_weight'] > 2800]

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                        '8' in data['locations'] and data['total_weight'] < 2800]

            if nedoves_in_location_8:
                min_num_th = min(
                    nedoves_in_location_8,
                    key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                    default=float('inf'))
                print("Num_th with maximum total_weight in nedoves_in_location_8:", min_num_th)

            pereves_8_code_tt = []

            if pereves_in_location_8:
                for num_th in pereves_in_location_8:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '8':
                                if total_weight > 2944:
                                    total_weight -= item['weight']
                                    pereves_8_code_tt.append(item['code_tt'])
                                else:
                                    break

            if pereves_8_code_tt:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(pereves_8_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
                session.commit()

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        print("End refactor_nedoves_num_th_8")
        return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_7_if_norm():
    print("Start refactor_7_if_norm")
    max_num_th = None
    min_num_th = None
    total_weight_min_th = None
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        pereves_in_location_7 = [num_th for num_th, data in total_weights_by_num_th if
                                    '7' in data['locations'] and data['total_weight'] > 3200]

        nedoves_in_location_7 = [num_th for num_th, data in total_weights_by_num_th if
                                    '7' in data['locations'] and data['total_weight'] < 2700]

        if nedoves_in_location_7:
            min_num_th = min(
                nedoves_in_location_7,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Минимальная в 9:", max_num_th)

        if min_num_th:
            if min_num_th in details_by_num_th:
                total_weight_min_th = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == min_num_th)

        if pereves_in_location_7:
            max_num_th = max(
                pereves_in_location_7,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("максимальная в 7:", max_num_th)

        pereves_code_tt_7 = []

        if max_num_th and min_num_th:
            if max_num_th in details_by_num_th:
                sorted_items = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'])
                total_weight = total_weight_min_th
                for item in sorted_items:
                    if item.get('location') == '7' or '6':
                        if total_weight <= 3150:
                            total_weight += item['weight']
                            pereves_code_tt_7.append(item['code_tt'])
                        else:
                            break

        if pereves_code_tt_7:
            session.query(AddressTable).filter(
                AddressTable.code_tt.in_(pereves_code_tt_7)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        print("End refactor_nedoves_num_th_8")
        return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def give_6_from_7_pereves_to_9_with_nedoves():
    print("Start refactor_7_pereves")
    max_num_th = None
    min_num_th = None
    total_weight_min_th = None
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        pereves_in_location_7 = [num_th for num_th, data in total_weights_by_num_th if
                                    '7' in data['locations'] and data['total_weight'] > 3200]

        nedoves_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if
                                    '9' in data['locations'] and data['total_weight'] < 2700]

        if nedoves_in_location_9:
            min_num_th = min(
                nedoves_in_location_9,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("Минимальная в 9:", max_num_th)

        if min_num_th:
            if min_num_th in details_by_num_th:
                total_weight_min_th = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == min_num_th)

        if pereves_in_location_7:
            max_num_th = max(
                pereves_in_location_7,
                key=lambda num_th: next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th),
                default=float('inf'))
            print("максимальная в 7:", max_num_th)

        pereves_code_tt_7 = []

        if max_num_th and min_num_th:
            if max_num_th in details_by_num_th:
                sorted_items = sorted(details_by_num_th[max_num_th], key=lambda x: x['weight'])
                total_weight = total_weight_min_th
                for item in sorted_items:
                    if item.get('location') == '6':
                        if total_weight <= 3150:
                            total_weight += item['weight']
                            pereves_code_tt_7.append(item['code_tt'])
                        else:
                            break

        if pereves_code_tt_7:
            session.query(AddressTable).filter(
                AddressTable.code_tt.in_(pereves_code_tt_7)).update({'num_th': min_num_th}, synchronize_session='fetch')
            session.commit()

        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        print("End refactor_nedoves_num_th_8")
        return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_8_pereves(result):
    print("Start refactor_8_pereves")
    try:
        if result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            pereves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                        '8' in data['locations'] and data['total_weight'] < 2800]

            location_8_6 = [num_th for num_th, data in total_weights_by_num_th if
                                '8' in data['locations'] and '6' in data['locations']]

            location_8_0 = [num_th for num_th, data in total_weights_by_num_th if
                                '8' in data['locations'] and '0' in data['locations']]

            location_6_code_tt = []

            location_0_code_tt = []

            if location_8_0:
                for num_th in location_8_0:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '0':
                                if total_weight != 0:
                                    total_weight -= item['weight']
                                    location_0_code_tt.append(item['code_tt'])
                                else:
                                    break

            if location_8_6:
                for num_th in location_8_6:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '0':
                                if total_weight != 0:
                                    total_weight -= item['weight']
                                    location_6_code_tt.append(item['code_tt'])
                                else:
                                    break

            if location_6_code_tt:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(location_6_code_tt)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            if location_0_code_tt:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(location_0_code_tt)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

        print("End refactor_nedoves_num_th_8")
        return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def refactor_8_6_and_8_0(result):
    print("Start refactor_nedoves_num_th_8")
    try:
        if not result:

            total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

            nedoves_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if
                                        '8' in data['locations'] and data['total_weight'] < 2800]

            location_8_6 = [num_th for num_th, data in total_weights_by_num_th if
                                '8' in data['locations'] and '6' in data['locations']]

            location_8_0 = [num_th for num_th, data in total_weights_by_num_th if
                                '8' in data['locations'] and '0' in data['locations']]

            location_6_code_tt = []

            location_0_code_tt = []

            if location_8_0:
                for num_th in location_8_0:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '0':
                                if total_weight != 0:
                                    total_weight -= item['weight']
                                    location_0_code_tt.append(item['code_tt'])
                                else:
                                    break

            if location_8_6:
                for num_th in location_8_6:
                    if num_th in details_by_num_th:
                        sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                        total_weight = next(
                            data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                        for item in sorted_items:
                            if item.get('location') == '0':
                                if total_weight != 0:
                                    total_weight -= item['weight']
                                    location_6_code_tt.append(item['code_tt'])
                                else:
                                    break

            if location_6_code_tt:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(location_6_code_tt)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

            if location_0_code_tt:
                session.query(AddressTable).filter(
                    AddressTable.code_tt.in_(location_0_code_tt)).update({'num_th': None}, synchronize_session='fetch')
                session.commit()

        print("End refactor_nedoves_num_th_8")
        return

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def check_weight_num_th_8(total_weights_by_num_th):
    print("Start check_weight_num_th_8")
    try:
        all_location_8 = [num_th for num_th, data in total_weights_by_num_th if '8' in data['locations']]

        total_weight_location_8 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_8)

        num_th_in_location_8 = len(all_location_8)

        if num_th_in_location_8 > 0:
            average_weight_location_8 = total_weight_location_8 - (3100 * num_th_in_location_8)
        else:
            average_weight_location_8 = 0.0

        print("Перевес на location 8:", average_weight_location_8)
        print("End check_weight_num_th_8")

        return 500 <= average_weight_location_8 <= 4000

    except ZeroDivisionError:
        return 0.0


def check_weight_num_th_7(total_weights_by_num_th):
    try:
        all_location_7 = [num_th for num_th, data in total_weights_by_num_th if '7' in data['locations']]

        total_weight_location_7 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_7)

        num_th_in_location_7 = len(all_location_7)

        if num_th_in_location_7 > 0:
            average_weight_location_7 = total_weight_location_7 - (3100 * num_th_in_location_7)
            if 500 <= average_weight_location_7 <= 4000:
                return True
            else:
                average_weight_location_7 = 0.0

        print(f"ТН с location 7 - {all_location_7}. Их текущий суммарный вес - {total_weight_location_7}. Их перевес - {average_weight_location_7}")

    except ZeroDivisionError:
        return 0.0


def check_weight_num_th_7():
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        all_location_7 = [num_th for num_th, data in total_weights_by_num_th if '7' in data['locations']]

        total_weight_location_7 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_7)

        num_th_in_location_7 = len(all_location_7)

        if num_th_in_location_7 > 0:
            average_weight_location_7 = total_weight_location_7 / num_th_in_location_7
        else:
            average_weight_location_7 = 0.0

        print(f"ТН с location 7 - {all_location_7}. Их текущий суммарный вес - {total_weight_location_7}."
              f" Их перевес - {average_weight_location_7}")

        if average_weight_location_7 >= 3300:
            return "Overweight"
        elif average_weight_location_7 < 2700:
            return "Underweight"
        else:
            return "Norm"

    except ZeroDivisionError:
        return "Error: Division by zero occurred"


def check_weight_num_th_6(total_weights_by_num_th):
    try:
        all_location_6 = [num_th for num_th, data in total_weights_by_num_th if '6' in data['locations']]

        total_weight_location_6 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_6)

        num_th_in_location_6 = len(all_location_6)

        if num_th_in_location_6 > 0:
            average_weight_location_6 = total_weight_location_6 - (3100 * num_th_in_location_6)
        else:
            average_weight_location_6 = 0.0

        print(f"ТН с location 6 - {all_location_6}. Их текущий суммарный вес - {total_weight_location_6}. Их перевес - {average_weight_location_6}")

        return 250 <= average_weight_location_6 <= 4000

    except ZeroDivisionError:
        return 0.0


def check_weight_num_th_5():
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        all_location_5 = [num_th for num_th, data in total_weights_by_num_th if '5' in data['locations']]

        total_weight_location_5 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_5)

        num_th_in_location_5 = len(all_location_5)

        if num_th_in_location_5 >= 2:
            average_weight_location_2 = total_weight_location_5 - (3000 * num_th_in_location_5)
            print(average_weight_location_2)
            return 400 <= average_weight_location_2 <= 4000

        if num_th_in_location_5 < 2:
            all_location_5_4 = [num_th for num_th, data in total_weights_by_num_th if
                                '4' in data['locations'] and '5' in data['locations']]
            all_location_4 = [num_th for num_th, data in total_weights_by_num_th if '4' in data['locations']]

            if all_location_5_4 and all_location_4:
                overweight_in_location_5 = [num_th for num_th, data in total_weights_by_num_th if
                                            '5' in data['locations'] and data['total_weight'] > 3100]

                overweight_code_tt_4_from_5 = []
                if overweight_in_location_5:
                    for num_th in overweight_in_location_5:
                        if num_th in details_by_num_th:
                            sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                            total_weight = next(
                                data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                            for item in sorted_items:
                                if item.get('location') == '4':
                                    if total_weight > 3300:
                                        total_weight -= item['weight']
                                        overweight_code_tt_4_from_5.append(item['code_tt'])
                                    else:
                                        break

                location_4_data = [(
                    num_th, data['total_weight']) for num_th, data in total_weights_by_num_th if '4' in data['locations']]
                min_num_th = min(location_4_data, key=lambda x: x[1])[0]
                if min_num_th:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(
                        overweight_code_tt_4_from_5)).update({'num_th': min_num_th}, synchronize_session='fetch')
                    session.commit()

                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == min_num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {min_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                    session.query(ReestrTable).filter(ReestrTable.num_th == min_num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')
                    session.commit()

                for num_th in overweight_in_location_5:
                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                        session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

    except ZeroDivisionError:
        return 0.0


# def check_weight_num_th_5(total_weights_by_num_th):
#     try:
#         all_location_5 = [num_th for num_th, data in total_weights_by_num_th if '5' in data['locations']]
#
#         total_weight_location_5 = sum(
#             data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_5)
#
#         num_th_in_location_5 = len(all_location_5)
#
#         if num_th_in_location_5 > 0:
#             average_weight_location_2 = total_weight_location_5 - (3000 * num_th_in_location_5)
#         else:
#             average_weight_location_2 = 0.0
#
#         print(average_weight_location_2)
#
#         return 400 <= average_weight_location_2 <= 4000
#
#     except ZeroDivisionError:
#         return 0.0


def check_weight_num_th_2():
    try:
        total_weights_by_num_th, details_by_num_th = calculate_details_by_num_th()

        #all_location_2 = [num_th for num_th, data in total_weights_by_num_th if '2' in data['locations']]
        all_location_2_and_1 = [num_th for num_th, data in total_weights_by_num_th if
                                '2' in data['locations'] or '1' in data['locations']]

        total_weight_location_2 = sum(
            data['total_weight'] for num_th, data in total_weights_by_num_th if num_th in all_location_2_and_1)

        num_th_in_location_2_1 = len(all_location_2_and_1)

        nedoves_in_location_2_1 = [num_th for num_th, data in total_weights_by_num_th
                                   if ('2' in data['locations'] or '1' in data['locations'])
                                   and data['total_weight'] < 2800]

        total_nedoves_in_location_2_1 = len(nedoves_in_location_2_1)

        if num_th_in_location_2_1 > 0:
            average_weight_location_2_1 = total_weight_location_2 - (3000 * num_th_in_location_2_1)
            print(average_weight_location_2_1)
            if 800 <= average_weight_location_2_1 <= 4000:
                return True

            elif total_nedoves_in_location_2_1 == 2:
                location_1_data = [(num_th, data['total_weight']) for num_th, data in total_weights_by_num_th if '1' in data['locations']]
                max_num_th = max(location_1_data, key=lambda x: x[1])[0]  # Получаем num_th с максимальным total_weight
                min_num_th = min(location_1_data, key=lambda x: x[1])[0]  # Получаем num_th с минимальным total_weight
                print(max_num_th, min_num_th)
                total_weight = next(
                    data['total_weight'] for num, data in total_weights_by_num_th if num == min_num_th)
                sorted_items = sorted(details_by_num_th[min_num_th], key=lambda x: x['weight'], reverse=True)

                from_min_to_max_code_tt = []

                for item in sorted_items:
                    if item.get('location') == '1' or '2':
                        if total_weight > 1700:
                            total_weight -= item['weight']
                            from_min_to_max_code_tt.append(item['code_tt'])
                        else:
                            break

                if from_min_to_max_code_tt:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(
                        from_min_to_max_code_tt)).update({'num_th': max_num_th}, synchronize_session='fetch')
                    session.commit()

                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == max_num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {max_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                        session.query(ReestrTable).filter(ReestrTable.num_th == max_num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()
                    else:
                        print("Error")
                    return

            else:
                overweight_in_location_2 = [num_th for num_th, data in total_weights_by_num_th if
                                        '2' in data['locations'] and data['total_weight'] > 3300]
                overweight_code_tt = []
                if overweight_in_location_2:
                    for num_th in overweight_in_location_2:
                        if num_th in details_by_num_th:
                            sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                            total_weight = next(
                                data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                            for item in sorted_items:
                                if item.get('location') == '1':
                                    if total_weight > 3300:
                                        total_weight -= item['weight']
                                        overweight_code_tt.append(item['code_tt'])
                                    else:
                                        break
                location_2_data = [(num_th, data['total_weight']) for num_th, data in total_weights_by_num_th if '2' in data['locations']]
                max_num_th = max(location_2_data, key=lambda x: x[1])[0]  # Получаем num_th с максимальным total_weight
                min_num_th = min(location_2_data, key=lambda x: x[1])[0]  # Получаем num_th с минимальным total_weight
                print(max_num_th, min_num_th)
                if max_num_th:
                    session.query(AddressTable).filter(AddressTable.code_tt.in_(
                        overweight_code_tt)).update({'num_th': min_num_th}, synchronize_session='fetch')
                    session.commit()

                total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                   func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                    .filter(AddressTable.num_th == min_num_th).group_by(AddressTable.num_th).first()

                if total_weight_query:
                    total_weight, total_count_boxes = total_weight_query
                    print(f"Total Weight for {min_num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")
                    session.query(ReestrTable).filter(ReestrTable.num_th == min_num_th).update(
                        {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                        synchronize_session='fetch')
                    session.commit()

                for num_th in overweight_in_location_2:
                    total_weight_query = session.query(func.sum(AddressTable.weight).label('total_weight'),
                                                       func.sum(AddressTable.count_boxes).label('total_count_boxes')) \
                        .filter(AddressTable.num_th == num_th).group_by(AddressTable.num_th).first()

                    if total_weight_query:
                        total_weight, total_count_boxes = total_weight_query
                        print(f"Total Weight for {num_th}: {total_weight}, Total Count Boxes: {total_count_boxes}")

                        session.query(ReestrTable).filter(ReestrTable.num_th == num_th).update(
                            {'total_weight': total_weight, 'total_count_boxes': total_count_boxes},
                            synchronize_session='fetch')
                        session.commit()

    except ZeroDivisionError:
        return 0.0

# def adjusting_location_9_? (total_weights_by_num_th, details_by_num_th):
#     try:
#         all_location_9 = {num_th: data['total_weight'] for num_th, data in total_weights_by_num_th if
#                           '9' in data['locations']}
#         overweight_in_location_9 = {num_th: data['total_weight'] for num_th, data in total_weights_by_num_th if
#                                     '9' in data['locations'] and data['total_weight'] > Decimal('3500')}
#
#         min_weight_num_th = min(all_location_9.items(), key=lambda x: x[1])
#         for num_th, total_weight in overweight_in_location_9.items():
#             if num_th in details_by_num_th:
#                 sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
#                 for item in sorted_items:
#                     if item.get('location') == '9':
#                         if total_weight > Decimal('3500'):
#                             total_weight -= item['weight']
#                             if min_weight_num_th[1] + item['weight'] <= Decimal('3500'):
#                                 min_weight_num_th = (min_weight_num_th[0], min_weight_num_th[1] + item['weight'])
#                                 session.query(AddressTable).filter_by(code_tt=item['code_tt']).update(
#                                     {'num_th': min_weight_num_th[0]}, synchronize_session='fetch')
#                             else:
#                                 all_location_9[min_weight_num_th[0]] = min_weight_num_th[1]
#                                 all_location_9[num_th] = total_weight
#                                 min_weight_num_th = min(
#                                     {k: v for k, v in all_location_9.items() if v <= Decimal('3500')}.items(),
#                                     key=lambda x: x[1])
#                         else:
#                             break
#         session.commit()
#         return "Location 9 adjusted successfully"
#
#     except Exception as e:
#         print(f"Error: {e}")
#         session.rollback()
#         return None


def adjusting_location_9(total_weights_by_num_th, details_by_num_th):
    try:
        all_location_9 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations']]
        overweight_in_location_9 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and data['total_weight'] > 3100]

        overweight_code_tt = []

        if overweight_in_location_9:
            for num_th in overweight_in_location_9:
                if num_th in details_by_num_th:
                    # Сортируем по весу от меньшего к большему
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                    total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                    for item in sorted_items:
                        if item.get('location') == '9':
                            if total_weight > 3100:
                                total_weight -= item['weight']
                                overweight_code_tt.append(item['code_tt'])
                            else:
                                break

        location_9_8 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '8' in data['locations']]
        location_9_7 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '7' in data['locations']]
        location_9_6 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '6' in data['locations']]
        if location_9_8:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_8), AddressTable.location == '8').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_8[0]}, synchronize_session='fetch')
            session.commit()
        if location_9_7:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_7), AddressTable.location == '7').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_7[0]}, synchronize_session='fetch')
            session.commit()
        if location_9_6:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_6), AddressTable.location == '6').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_6[0]}, synchronize_session='fetch')
            session.commit()

    except Exception as e:
        print(f"Error: {e}")


def adjusting_location_8(total_weights_by_num_th, details_by_num_th):
    try:
        all_location_8 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations']]
        overweight_in_location_8 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and data['total_weight'] > 3500]

        overweight_code_tt = []

        if overweight_in_location_8:
            for num_th in overweight_in_location_8:
                if num_th in details_by_num_th:
                    # Сортируем по весу от меньшего к большему
                    sorted_items = sorted(details_by_num_th[num_th], key=lambda x: x['weight'])
                    total_weight = next(data['total_weight'] for num, data in total_weights_by_num_th if num == num_th)

                    for item in sorted_items:
                        if item.get('location') == '9':
                            if total_weight > 3500:
                                total_weight -= item['weight']
                                overweight_code_tt.append(item['code_tt'])
                            else:
                                break

        location_9_8 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '8' in data['locations']]
        location_9_7 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '7' in data['locations']]
        location_9_6 = [num_th for num_th, data in total_weights_by_num_th if '9' in data['locations'] and '6' in data['locations']]
        if location_9_8:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_8), AddressTable.location == '8').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_8[0]}, synchronize_session='fetch')
            session.commit()
        if location_9_7:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_7), AddressTable.location == '7').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_7[0]}, synchronize_session='fetch')
            session.commit()
        if location_9_6:
            session.query(AddressTable).filter(AddressTable.num_th.in_(location_9_6), AddressTable.location == '6').update({'num_th': None}, synchronize_session='fetch')
            session.query(AddressTable).filter(AddressTable.code_tt.in_(overweight_code_tt)).update({'num_th': location_9_6[0]}, synchronize_session='fetch')
            session.commit()

    except Exception as e:
        print(f"Error: {e}")



# def can_assign_driver(current_location, current_weight, location, priority, weight):
#     weight_limit = 3000
#     if current_weight + weight > weight_limit:
#         return False
#
#     location = int(location)
#
#     location_rules = {
#         7: (6, 50),
#         6: (5, 50),
#         5: (4, 50),
#         4: (3, 50),
#         3: (2, 50),
#         2: (1, 50),
#         1: (0, 50)
#     }
#
#     if current_location in location_rules:
#         next_location, max_priority = location_rules[current_location]
#         if location == next_location and priority < max_priority:
#             return True
#
#     return int(current_location) < int(location)


# def assign_drivers_to_addresses():
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#     driver_locations = defaultdict(int)
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#
#     drivers = session.query(DriversTable).all()
#     driver_list = {driver.last_name: driver for driver in drivers}
#
#     for details in sorted(details_by_code_tt, key=lambda x: (x['location'], -x['priority']), reverse=True):
#         code_tt = details['code_tt']
#         weight = float(details['weight'])
#         location = details['location']
#         priority = details['priority']
#
#         assigned_driver = None
#         for driver_name, driver in driver_list.items():
#             current_weight = driver_weights[driver_name]
#             current_location = driver_locations[driver_name]
#
#             if can_assign_driver(current_location, current_weight, location, priority, weight):
#                 assigned_driver = driver_name
#                 driver_weights[driver_name] += weight
#                 driver_locations[driver_name] = location
#                 break
#
#         if not assigned_driver:
#             assigned_driver = 'Нет доступных водителей'
#
#         driver_assignments[code_tt] = assigned_driver
#
#         address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
#         if address_record:
#             address_record.last_name = assigned_driver if assigned_driver != 'Нет доступных водителей' else None
#             session.commit()
#
#         current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
#         print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, {assigned_driver}, {current_load} /3000")
#
#     return driver_assignments


# def assign_drivers_to_addresses():
#     print("Start assign_drivers_to_addresses")
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#     driver_locations = defaultdict(int)
#     assigned_tasks = set()
#
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#
#     drivers = session.query(DriversTable).all()
#     driver_list = {driver.last_name: {"driver_obj": driver, "location": driver.location} for driver in drivers}
#
#     sorted_driver_list = sorted(driver_list.items(), key=lambda x: x[1]['location'], reverse=True)
#
#     for driver_name, driver_info in sorted_driver_list:
#         current_location = driver_info["location"]
#         current_weight = driver_weights[driver_name]
#
#         for details in sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True):
#             code_tt = details['code_tt']
#             weight = float(details['weight'])
#             location = details['location']
#             priority = details['priority']
#
#             if code_tt in assigned_tasks:  # Проверяем, назначено ли уже задание
#                 continue
#
#             if current_weight + weight <= 3000 and abs(current_location - location) <= 1:
#                 driver_weights[driver_name] += weight
#                 driver_locations[driver_name] = location
#                 driver_assignments[code_tt] = driver_name
#                 assigned_tasks.add(code_tt)
#
#                 address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
#                 if address_record:
#                     address_record.last_name = driver_name
#                     session.commit()
#
#                 print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, Водитель: {driver_name}, Общий вес: {driver_weights[driver_name]} /3000")
#
#             if current_weight + weight >= 3000:
#                 break
#
#     return driver_assignments
from sqlalchemy import desc


# def assign_drivers_to_addresses():
#     print("Start assign_drivers_to_addresses")
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#     driver_locations = defaultdict(int)
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#     details_by_code_tt = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     drivers = session.query(DriversTable).order_by(desc(DriversTable.location)).all()
#     driver_list = {driver.last_name: {"driver_obj": driver, "location": driver.location} for driver in drivers}
#
#     for driver in drivers:
#         driver_locations[driver.last_name] = driver.location
#
#     for details in details_by_code_tt:
#         code_tt = details['code_tt']
#         weight = float(details['weight'])
#         location = details['location']
#         priority = details['priority']
#
#         assigned_driver = None
#         for driver_name, driver in driver_list.items():
#             print(f"Проверка задания {code_tt} для водителя {driver_name}")
#             current_weight = driver_weights[driver_name]
#             current_location = driver_locations[driver_name]
#             print(f"Текущее местоположение {current_location} и местоположение {location}. Текущий вес {current_weight}, новый вес {weight}, общий вес {current_weight + weight}")
#
#             if current_weight + weight <= 3000:
#                 driver_loc = driver['location']
#
#                 if driver_loc == 7 and not (location == 7 or (location == 6 and priority < 50)):
#                     continue
#
#                 if driver_loc == 6 and not (location == 6 or (location == 5 and priority < 50)):
#                     continue
#
#                 if driver_loc == 5 and not (location == 5 or (location == 4 and priority < 50)):
#                     continue
#
#                 if driver_loc == 4 and not (location == 4 or (location == 3 and priority < 50)):
#                     continue
#
#                 if driver_loc == 3 and not (location == 3 or (location == 2 and priority < 50)):
#                     continue
#
#                 if driver_loc == 2 and not (location == 2 or (location == 1 and priority < 50)):
#                     continue
#
#                 if driver_loc == 1 and not (location == 1 or (location == 0 and priority < 50)):
#                     continue
#
#                 assigned_driver = driver_name
#                 if not assigned_driver:
#                     continue
#
#             if assigned_driver:
#                 driver_weights[driver_name] += weight
#                 driver_locations[driver_name] = location
#                 break
#
#         driver_assignments[code_tt] = assigned_driver
#
#         address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
#         if address_record:
#             address_record.last_name = assigned_driver if assigned_driver != 'Нет доступных водителей' else None
#             session.commit()
#
#         current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
#         print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, {assigned_driver}, {current_load} /3000")
#
#     return driver_assignments


from collections import defaultdict
from sqlalchemy import desc


#
#
# def assign_drivers_to_addresses():
#     print("Start assign_drivers_to_addresses")
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#     driver_locations = defaultdict(int)
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#     details_by_code_tt = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     drivers = session.query(DriversTable).order_by(desc(DriversTable.location)).all()
#     driver_list = {driver.last_name: driver.location for driver in drivers}
#
#     for details in details_by_code_tt:
#         code_tt = details['code_tt']
#         weight = float(details['weight'])
#         location = details['location']
#         priority = details['priority']
#
#         assigned_driver = None
#         for driver_name, driver_location in driver_list.items():
#             print(f"Проверка задания {code_tt} для водителя {driver_name}")
#             current_weight = driver_weights[driver_name]
#             print(
#                 f"Текущее местоположение {driver_location} и местоположение {location}. Текущий вес {current_weight}, новый вес {weight}, общий вес {current_weight + weight}")
#
#             if current_weight + weight <= 3000 and driver_location == location:
#                 assigned_driver = driver_name
#                 break
#
#         if assigned_driver:
#             driver_weights[assigned_driver] += weight
#             driver_locations[assigned_driver] = location
#
#         driver_assignments[code_tt] = assigned_driver
#
#         address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
#         if address_record:
#             address_record.last_name = assigned_driver if assigned_driver != 'Нет доступных водителей' else None
#             session.commit()
#
#         current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
#         print(
#             f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, {assigned_driver}, {current_load} /3000")


def assign_drivers_to_addresses():
    print("Start assign_drivers_to_addresses")
    driver_assignments = {}
    driver_weights = defaultdict(float)
    driver_locations = defaultdict(int)

    details_by_code_tt = calculate_weights_and_get_driver_last_names()

    drivers = session.query(DriversTable).order_by(desc(DriversTable.location)).all()
    driver_list = {driver.last_name: [driver.location, 0.0] for driver in drivers}

    drivers_all = {}
    for driver in drivers:
        if drivers_all.get(driver.location) is None:
            drivers_all[driver.location] = {driver.last_name: {"weight": 0.0, "addreses": []}}
        else:
            drivers_all[driver.location][driver.last_name] = {"weight": 0.0, "addreses": []}

    for details in sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True):
        code_tt = details['code_tt']
        weight = float(details['weight'])
        location = details['location']
        priority = details['priority']

        assigned_driver = None
        if location == '0':
            drivers_on_1_2 = {}
            drivers_on_1_2.update(drivers_all[1])
            drivers_on_1_2.update(drivers_all[2])
            for driver_on_location in drivers_on_1_2.keys():
                if drivers_on_1_2[driver_on_location]["weight"] + weight <= 3200.0 and drivers_on_1_2[driver_on_location]["weight"] != 0.0:
                    assigned_driver = driver_on_location
                    for location_find in drivers_all.keys():
                        for driver_in_location in drivers_all[location_find].keys():
                            if driver_in_location == driver_on_location:
                                drivers_all[location_find][driver_in_location]["weight"] += weight
                                drivers_all[location_find][driver_in_location]["addreses"].append(details)
                    break
        else:
            if priority < 50 and int(location) != len(drivers_all.keys()):
                drivers_on_location = drivers_all[int(location)+1]
                for driver_on_location in drivers_on_location.keys():
                    if drivers_on_location[driver_on_location]["weight"] + weight <= 3200.0:
                        assigned_driver = driver_on_location
                        drivers_on_location[driver_on_location]["weight"] += weight
                        drivers_on_location[driver_on_location]["addreses"].append(details)
                        break
            else:
                drivers_on_location = drivers_all[int(location)]
                for driver_on_location in drivers_on_location.keys():
                    if drivers_on_location[driver_on_location]["weight"] + weight <= 3200.0:
                        assigned_driver = driver_on_location
                        drivers_on_location[driver_on_location]["weight"] += weight
                        drivers_on_location[driver_on_location]["addreses"].append(details)
                        break
        if not assigned_driver:
            print(f"Водители по локации {location} загружены полностью")

        # for driver_name, driver_location in driver_list.items():
        #     print(f"Проверка задания {code_tt} для водителя {driver_name}")
        #     current_weight = driver_weights[driver_name]
        #     print(f"Текущее местоположение {driver_location} и местоположение {location}. Текущий вес {current_weight}, новый вес {weight}, общий вес {current_weight + weight}")
        #
        #     if current_weight + weight <= 3200:
        #         if driver_location == 10 and int(location) == 10:
        #             assigned_driver = driver_name
        #         elif driver_location == 10 and int(location) == 9 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 9 and int(location) == 9:
        #             assigned_driver = driver_name
        #         elif driver_location == 9 and int(location) == 8 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 8 and int(location) == 8:
        #             assigned_driver = driver_name
        #         elif driver_location == 8 and int(location) == 7 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 7 and int(location) == 7:
        #             assigned_driver = driver_name
        #         elif driver_location == 7 and int(location) == 6 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 6 and int(location) == 6:
        #             assigned_driver = driver_name
        #         elif driver_location == 6 and int(location) == 5 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 5 and int(location) == 5:
        #             assigned_driver = driver_name
        #         elif driver_location == 5 and int(location) == 4 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 4 and int(location) == 4:
        #             assigned_driver = driver_name
        #         elif driver_location == 4 and int(location) == 5 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 4 and int(location) == 3 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 3 and int(location) == 3:
        #             assigned_driver = driver_name
        #         elif driver_location == 3 and int(location) == 2 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 3 and int(location) == 0 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 2 and int(location) == 2:
        #             assigned_driver = driver_name
        #         elif driver_location == 2 and int(location) == 1 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 2 and int(location) == 0 and priority < 50:
        #             assigned_driver = driver_name
        #         elif driver_location == 1 and int(location) == 1:
        #             assigned_driver = driver_name
        #         elif driver_location == 1 and int(location) == 0:
        #             assigned_driver = driver_name
        #
        #     if assigned_driver:
        #         driver_weights[driver_name] += weight
        #         driver_locations[driver_name] = location
        #         break

        driver_assignments[code_tt] = assigned_driver

        address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
        if address_record:
            address_record.last_name = assigned_driver if assigned_driver != 'Нет доступных водителей' else None
            session.commit()

        current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
        print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, {assigned_driver}, {current_load} / 3000")
    # for location_driver in drivers_all.keys():
    #     if location_driver != 1:
    #         for drivers_with_weight_now in drivers_all[location_driver].keys():
    #             if drivers_all[location_driver][drivers_with_weight_now]["weight"] > 3000.0:
    #                 continue
    #             else:
    #                 for drivers_with_weight_low in drivers_all[location_driver - 1].keys():
    #                     for address in drivers_all[location_driver][drivers_with_weight_low]['addreses']:
    #                         if address['priority'] < 50:
    #                             drivers_all[location_driver][drivers_with_weight_now]['addreses'].append(address)
    #                             drivers_all[location_driver][drivers_with_weight_low]['addreses'].remove(address)

    return driver_assignments


# def calculate_weights_and_get_driver_last_names():
#     try:
#         total_weight_tomorrow = 0.0
#         weights_by_location = defaultdict(float)
#         tomorrow = datetime.now().date() + timedelta(days=1)
#         details_by_code_tt = []
#
#         addresses_tomorrow = session.query(AddressTable).filter(AddressTable.date_th == tomorrow).all()
#
#         for address in addresses_tomorrow:
#             address_weight = float(address.weight.real)
#             total_weight_tomorrow += address_weight
#
#             first_digit = next((char for char in address.location if char.isdigit()), None)
#             if first_digit:
#                 rounded_weight = math.ceil(address_weight)
#                 weights_by_location[first_digit] += rounded_weight
#
#             detail = {
#                 'code_tt': address.code_tt,
#                 'weight': address.weight,
#                 'location': address.location,
#                 'priority': address.priority
#             }
#             details_by_code_tt.append(detail)
#
#         total_batches = math.ceil(total_weight_tomorrow / 1000)
#         logger.info(f"Total weight for tomorrow: {total_weight_tomorrow} kg, requiring {total_batches} drivers")
#
#         for location, weight in weights_by_location.items():
#             logger.info(f"Total weight for location {location}: {int(weight)} kg")
#
#         return details_by_code_tt
#     except Exception as e:
#         logger.error(f"Error in calculate_weights_and_get_driver_last_names: {e}")
#         return None
#
#
# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#
#     sorted_details = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     drivers = session.query(DriversTable).all()
#     for driver in drivers:
#         drivers_by_location[driver.location].append(driver.last_name)
#
#     for details in sorted_details:
#         location = details['location']
#         priority = details['priority']
#         code_tt = details['code_tt']
#         weight = float(details['weight'])
#
#         assigned_driver = None
#         for driver in drivers_by_location[location]:
#             if driver_weights[driver] + weight <= 1000:
#                 assigned_driver = driver
#                 driver_weights[driver] += weight
#                 break
#
#         if not assigned_driver and drivers_by_location[location]:
#             new_driver = drivers_by_location[location].pop(0)
#             assigned_driver = new_driver
#             driver_weights[new_driver] = weight
#
#         if not assigned_driver:
#             assigned_driver = 'Нет доступных водителей'
#
#         driver_assignments[code_tt] = assigned_driver
#
#         address_record = session.query(AddressTable).filter_by(code_tt=code_tt).first()
#         if address_record:
#             address_record.last_name = assigned_driver if assigned_driver != 'Нет доступных водителей' else None
#             session.commit()
#
#         current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
#         print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {details['priority']}, {assigned_driver}, {current_load} /1000")
#
#     return driver_assignments
#
#
# def calculate_driver_loads():
#     driver_loads = defaultdict(Decimal)
#     address_records = session.query(AddressTable).all()
#
#     for record in address_records:
#         if record.last_name:
#             driver_loads[record.last_name] += Decimal(record.weight)
#
#     return driver_loads


# def reassign_tasks_to_drivers_with_capacity():
#     driver_loads = calculate_driver_loads()
#     drivers = session.query(DriversTable).all()
#     drivers_with_capacity = {driver.last_name: driver_loads[driver.last_name] for driver in drivers if driver_loads[driver.last_name] < 1000}
#
#     for driver_name, load in drivers_with_capacity.items():
#         driver = session.query(DriversTable).filter_by(last_name=driver_name).first()
#         if driver:
#             driver_location = str(int(driver.location) - 1) if driver.location.isdigit() else None
#             if driver_location is not None:
#                 eligible_tasks = session.query(AddressTable)\
#                                         .filter(AddressTable.location == driver_location, AddressTable.priority <= 50)\
#                                         .all()
#
#                 for task in eligible_tasks:
#                     if load + task.weight <= 1000:
#                         task.last_name = driver_name
#                         load += task.weight
#
#                         driver_loads[driver_name] = load
#
#     session.commit()


def display_assigned_driver_details():
    address_records = session.query(AddressTable).all()
    driver_details = defaultdict(lambda: {'total_weight': 0, 'assignments': []})

    for record in address_records:
        assigned_driver = record.last_name if record.last_name else 'Нет доступных водителей'
        weight = record.weight
        code_tt = record.code_tt
        location = record.location
        priority = record.priority

        driver_details[assigned_driver]['total_weight'] += weight
        driver_details[assigned_driver]['assignments'].append({
            'code_tt': code_tt,
            'location': location,
            'priority': priority,
            'weight': weight
        })

    for driver, details in driver_details.items():
        print(f"Водитель: {driver}, Общий вес: {details['total_weight']} /3000")
        for assignment in details['assignments']:
            print(f"  Code_tt: {assignment['code_tt']}, Местоположение: {assignment['location']}, Приоритет: {assignment['priority']}, Вес: {assignment['weight']}")




# def assign_drivers_to_addresses():
#     driver_assignments = {}
#     driver_weights = defaultdict(float)
#     driver_lowest_location = defaultdict(lambda: float('inf'))
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#     sorted_details = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     drivers = session.query(DriversTable).all()
#     available_drivers = [driver.last_name for driver in drivers]
#
#     for details in sorted_details:
#         location = details['location']
#         code_tt = details['code_tt']
#         weight = float(details['weight'])
#         priority = details['priority']
#
#         assigned_driver = None
#         for driver in available_drivers:
#             if driver_weights[driver] + weight <= 1000:
#                 assigned_driver = driver
#                 driver_weights[driver] += weight
#                 break
#
#         if not assigned_driver and priority <= 50:
#             for lower_location_driver in available_drivers:
#                 if driver_weights[lower_location_driver] + weight <= 1000:
#                     assigned_driver = lower_location_driver
#                     driver_weights[lower_location_driver] += weight
#                     break
#
#         if not assigned_driver:
#             assigned_driver = 'Нет доступных водителей'
#
#         current_load = driver_weights[assigned_driver] if assigned_driver != 'Нет доступных водителей' else 0
#         print(f"Code_tt: {code_tt}, Вес: {weight}, Местоположение: {location}, Приоритет: {priority}, {assigned_driver}, {current_load} /1000")
#
#     return driver_assignments









# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     driver_assignments = {}
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#
#     sorted_details = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     with Session() as db:
#         drivers = db.query(DriversTable).all()
#         for driver in drivers:
#             drivers_by_location[driver.location].append(driver.last_name)
#
#     for details in sorted_details:
#         location = details['location']
#         code_tt = details['code_tt']
#
#         if drivers_by_location[location]:
#             driver_last_name = drivers_by_location[location].pop(0)
#             driver_assignments[code_tt] = driver_last_name
#         else:
#             driver_assignments[code_tt] = 'Нет доступных водителей'
#
#         print(f"Code_tt: {code_tt}, Вес: {details['weight']}, Местоположение: {details['location']}, Приоритет: {details['priority']}, Водитель: {driver_assignments[code_tt]}")
#
#     return driver_assignments



# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     driver_assignments = {}
#
#     details_by_code_tt = calculate_weights_and_get_driver_last_names()
#
#     sorted_details = sorted(details_by_code_tt, key=lambda x: (x['location'], x['priority']), reverse=True)
#
#     for details in sorted_details:
#         print(f"Code_tt: {details['code_tt']}, Вес: {details['weight']}, Местоположение: {details['location']}, Приоритет: {details['priority']}")
#
#     with Session() as db:
#         drivers = db.query(DriversTable).all()
#         for driver in drivers:
#             drivers_by_location[driver.location].append(driver.last_name)
#
#     for details in sorted_details:
#         location = details['location']
#         code_tt = details['code_tt']
#
#         if drivers_by_location[location]:
#             driver_last_name = drivers_by_location[location].pop(0)
#             driver_assignments[code_tt] = driver_last_name
#         else:
#             driver_assignments[code_tt] = 'Нет доступных водителей'
#
#     for code_tt, driver in driver_assignments.items():
#         print(f"Code_tt {code_tt}: Назначен водитель {driver}")
#
#     return driver_assignments





# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     location_driver_assignments = defaultdict(list)
#     weights_by_location = calculate_weights_and_get_driver_last_names()
#     print(f" weights_by_location: {weights_by_location}")
#
#     with Session() as db:
#         drivers = db.query(DriversTable).all()
#         for driver in drivers:
#             drivers_by_location[driver.location].append(driver.last_name)
#
#     for location, weight in weights_by_location.items():
#         num_drivers_needed = math.ceil(weight / 1000)
#         available_drivers = drivers_by_location[location]
#
#         while num_drivers_needed > 0 and available_drivers:
#             driver_last_name = available_drivers.pop(0)
#             location_driver_assignments[location].append(driver_last_name)
#             num_drivers_needed -= 1
#
#     for location, drivers in location_driver_assignments.items():
#         for driver in drivers:
#             print(f"Локация {location}: Назначен водитель {driver}")
#
#     return location_driver_assignments



# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     location_driver_assignments = defaultdict(list)
#     weights_by_location = calculate_weights_and_get_driver_last_names()
#
#     with Session() as db:
#         drivers = db.query(DriversTable).all()
#         for driver in drivers:
#             drivers_by_location[driver.location].append(driver.last_name)
#
#     tomorrow = datetime.now().date() + timedelta(days=1)
#     with db:
#         addresses_tomorrow = db.query(AddressTable).filter(AddressTable.date_th == tomorrow).all()
#         for address in addresses_tomorrow:
#             first_digit = next((char for char in address.location if char.isdigit()), None)
#             if first_digit:
#                 location_weight = weights_by_location.get(first_digit, 0)
#                 num_drivers_needed = math.ceil(location_weight / 3000)
#                 available_drivers = drivers_by_location[first_digit]
#
#                 while num_drivers_needed > 0 and available_drivers:
#                     driver_last_name = available_drivers.pop(0)
#                     location_driver_assignments[first_digit].append(driver_last_name)
#                     num_drivers_needed -= 1
#
#     for location, drivers in location_driver_assignments.items():
#         for driver in drivers:
#             print(f"Локация {location}: Назначен водитель {driver}")
#
#     return location_driver_assignments


# def assign_drivers_to_addresses():
#     drivers_by_location = defaultdict(list)
#     assigned_locations = set()
#
#     with Session() as db:
#         drivers = db.query(DriversTable).all()
#         for driver in drivers:
#             drivers_by_location[driver.location].append(driver.last_name)
#
#     tomorrow = datetime.now().date() + timedelta(days=1)
#     with db:
#         addresses_tomorrow = db.query(AddressTable).filter(AddressTable.date_th == tomorrow).all()
#         for address in addresses_tomorrow:
#             first_digit = next((char for char in address.location if char.isdigit()), None)
#             if first_digit and drivers_by_location[first_digit] and first_digit not in assigned_locations:
#                 driver_last_name = drivers_by_location[first_digit][0]
#                 address.last_name = driver_last_name
#                 print(f"Локация {address.location}: Назначен водитель {driver_last_name}")
#                 assigned_locations.add(first_digit)
#
#     return drivers_by_location







# def overweight_light():
#     try:
#         with Session() as db:
#             num_ths = db.query(AddressTable.num_th).distinct().all()
#             for num_th in num_ths:
#                 addresses = db.query(
#                     AddressTable).filter_by(num_th=num_th[0]).order_by(AddressTable.index_number.asc()).all()
#                 weight = 0.0
#                 for address in addresses:
#                     weight += float(address.weight.real)
#                     if weight > 3000.0:
#                         address.last_name = address.last_name + " - перевес"  # Помечаем последующие адреса группы num_th как "перевес"
#
#                 db.commit()
#     except Exception as e:
#         print(str(e))

