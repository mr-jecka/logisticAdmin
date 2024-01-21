import tempfile
import time
from copy import copy

from openpyxl import Workbook
from aiogram import Bot, Dispatcher, executor, types
import markup as nav
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from database import get_drivers_for_route, get_routes, insert_driver_for_route, get_optimal_json, \
    get_main_json, get_info_for_report, insert_user_id_for_addresses, get_main_json_14, calculate_weights_by_num_th,\
    display_assigned_driver_details, assign_drivers_to_addresses, num_th_to_drivers, update_index_numbers
from aiogram.types import CallbackQuery, KeyboardButton, ReplyKeyboardMarkup
from aiogram.dispatcher import FSMContext
import logging
from aiogram import types
from PIL import Image, ImageDraw, ImageFont
import database
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%d-%b-%y %H:%M:%S'
)
logger = logging.getLogger(__name__)

TOKEN = "6441679596:AAEYabzPiA4dg0GOlBISJk0BhAjqn1OPjF0"
# TOKEN = "6441679596:AAEYabzPiA4dg0GOlBISJk0BhAjqn1OPjF0"
#TOKEN = "6489569901:AAHBPmIvgYsxj_M_p6x9FnG_RThYEBthcRc" #@MoveTrafficBot

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)


class EnterForm(StatesGroup):
    waiting_for_scheduled_arrival = State()
    waiting_for_actual_arrival = State()
    waiting_for_shipment = State()
    waiting_for_departure = State()
    waiting_for_reestr = State()
    #waiting_for_routes = State()
    #waiting_for_driver = State()


nav.init(dp)


@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    mess = f'👋  Приветствую, {message.from_user.first_name}!  Я помощник по логистике! '
    await bot.send_message(message.chat.id, mess, reply_markup=nav.mainMenu)
    logging.info(f"User {message.from_user.username} {message.from_user.id} started the bot.")


@dp.callback_query_handler(text="Cancel")
async def cancel_handler(callback_query: types.CallbackQuery):
    await bot.send_message(callback_query.from_user.id, "Галя, у нас отмена!")
    await start(callback_query.message)


def update_excel_with_route_driver_car(selected_route, selected_driver_name, selected_driver_car, selected_time):
    wb = openpyxl.load_workbook('reestr.xlsx')
    sheet = wb.active

    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value == selected_route:
            sheet.cell(row=row, column=12).value = selected_driver_name
            sheet.cell(row=row, column=13).value = selected_time
            sheet.cell(row=row, column=14).value = selected_driver_car
            break

    wb.save('reestr.xlsx')


class RouteSelection(StatesGroup):
    waiting_for_route = State()
    waiting_for_driver = State()
    waiting_for_time = State()


@dp.callback_query_handler(text="optimalRoute")
async def start_distribute_route(query: types.CallbackQuery):
    await bot.send_message(query.from_user.id, "Строим оптимальные маршруты на завтра")
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    tomorrow_routes = update_index_numbers()
    if not tomorrow_routes:
        await bot.send_message(query.from_user.id, "Не найдены маршруты на завтра")
        return


@dp.callback_query_handler(text="overWeight")
async def start_distribute_route(query: types.CallbackQuery):
    assign_drivers_to_addresses()
    time.sleep(2)
    display_assigned_driver_details()
    #reassign_tasks_to_drivers_with_capacity()
    return


@dp.callback_query_handler(text="overWeight_2")
async def start_distribute_route(query: types.CallbackQuery):
    calculate_weights_by_num_th()
    #assign_drivers_to_addresses()
    time.sleep(2)
    #display_assigned_driver_details()
    #reassign_tasks_to_drivers_with_capacity()
    return


@dp.callback_query_handler(text="THforDrivers")
async def start_distribute_route(query: types.CallbackQuery):
    num_th_to_drivers()


@dp.callback_query_handler(text="distribute_routes")
async def start_distribute_route(query: types.CallbackQuery):
    await bot.send_message(
        query.from_user.id, "Вы хотите распределить маршруты на сегодня или на завтра ?", reply_markup=nav.mainMenu2)


@dp.callback_query_handler(text="distr_tomorrow")
async def start_distribute_tomorrow_routes(query: types.CallbackQuery):
    await bot.send_message(query.from_user.id, "Распределение маршрутов на завтра...")
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    tomorrow_routes = get_routes(tomorrow)
    if not tomorrow_routes:
        await bot.send_message(query.from_user.id, "Не найдены маршруты на завтра")
        return

    route_buttons = [KeyboardButton(route[0]) for route in tomorrow_routes]
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(*route_buttons)
    await bot.send_message(query.from_user.id, "Выберете маршрут:", reply_markup=keyboard)
    await RouteSelection.waiting_for_route.set()


@dp.callback_query_handler(text="distr_today")
async def start_distribute_today_routes(query: types.CallbackQuery):
    await bot.send_message(query.from_user.id, "Распределение маршрутов на сегодня...")
    today = datetime.now().date()
    today_routes = get_routes(today)
    if not today_routes:
        await bot.send_message(query.from_user.id, "Не найдены маршруты на сегодня")
        return

    route_buttons = [KeyboardButton(route[0]) for route in today_routes]
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(*route_buttons)
    await bot.send_message(query.from_user.id, "Выберете маршрут:", reply_markup=keyboard)

    await RouteSelection.waiting_for_route.set()


@dp.message_handler(state=RouteSelection.waiting_for_route)
async def handle_route_choice(message: types.Message, state: FSMContext):
    selected_route = message.text
    await state.update_data(selected_route=selected_route)

    all_drivers = get_drivers_for_route()

    if not all_drivers:
        await bot.send_message(message.from_user.id, "Водители не найдены")
        return

    driver_buttons = [KeyboardButton(driver[0]) for driver in all_drivers]
    driver_keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(*driver_buttons)
    await bot.send_message(message.from_user.id, "Выберете водителя:", reply_markup=driver_keyboard)
    await RouteSelection.waiting_for_driver.set()


@dp.message_handler(state=RouteSelection.waiting_for_driver)
async def handle_driver_choice(message: types.Message, state: FSMContext):
    selected_driver = message.text
    await state.update_data(selected_driver=selected_driver)
    loading_times = ["4:00", "4:30", "5:00", "5:30", "6:00", "6:30", "7:00", "7:30", "8:00"]
    time_buttons = [KeyboardButton(time) for time in loading_times]
    time_keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(*time_buttons)
    await bot.send_message(message.from_user.id, "Какое время подачи на погрузку?", reply_markup=time_keyboard)
    await RouteSelection.waiting_for_time.set()


@dp.message_handler(state=RouteSelection.waiting_for_time)
async def handle_loading_time_choice(message: types.Message, state: FSMContext):
    selected_time = message.text
    logger.info(f"Выбранное время: {selected_time}")

    user_data = await state.get_data()
    selected_route = user_data['selected_route']
    selected_driver = user_data['selected_driver']

    logger.info(f"Извлеченные данные из состояния: маршрут - {selected_route}, водитель - {selected_driver}")

    all_drivers = get_drivers_for_route()
    logger.info(f"Получен список водителей: {all_drivers}")

    driver_to_car = {driver[0]: driver[1] for driver in all_drivers}
    driver_to_user_id = {driver[0]: driver[2] for driver in all_drivers}

    selected_driver_car = driver_to_car.get(selected_driver)
    selected_driver_user_id = driver_to_user_id.get(selected_driver)

    if not selected_driver_car or not selected_driver_user_id:
        logger.error("Не найдена информация о машине или user_id водителя.")
        await bot.send_message(message.from_user.id, "Информация о водителе или его машине не найдена")
        return

    logger.info(f"Выбранный автомобиль водителя: {selected_driver_car}")
    logger.info(f"User ID выбранного водителя: {selected_driver_user_id}")

    try:
        insert_driver_for_route(
            selected_route, selected_driver, selected_driver_car, selected_time, selected_driver_user_id)
        logger.info("Водитель добавлен к маршруту успешно")
    except Exception as e:
        logger.exception("Ошибка при добавлении водителя к маршруту:")
        await bot.send_message(message.from_user.id, "Произошла ошибка при добавлении водителя к маршруту")
        return

    try:
        insert_user_id_for_addresses(selected_route, selected_driver_user_id)  # Обновление user_id в адресной таблице
        logger.info(f"User ID водителя обновлен для адресов маршрута {selected_route}.")
    except Exception as e:
        logger.exception("Ошибка при обновлении user_id в адресной таблице:")
        await bot.send_message(message.from_user.id, "Произошла ошибка при обновлении user_id в адресах")
        return

    await bot.send_message(
        message.from_user.id,
        f"Маршруту {selected_route} присвоен водитель {selected_driver} с машиной {selected_driver_car} и временем прибытия {selected_time}")
    await bot.send_message(message.from_user.id, "Продолжим распределение маршрутов ?", reply_markup=nav.mainMenu2)

    logger.info("Сообщение пользователю отправлено")
    await state.finish()


@dp.message_handler(lambda message: message.text.isdigit())
async def distribute_driver(message: types.Message, state: FSMContext):
    try:
        logging.info(f"Start distribute_route")
        route_number = message.text
        await state.update_data(selected_route=route_number)
        drivers = get_available_drivers()
        if drivers:
            driver_buttons = [KeyboardButton(driver[0]) for driver in drivers]
            keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(*driver_buttons)
            await message.answer("Select a driver:", reply_markup=keyboard)
        else:
            await message.answer("No available drivers.")
    except Exception as e:
        logging.error(f"Error in distribute_route: {str(e)}")


@dp.callback_query_handler(text="inputBD")
async def address(message: types.Message):

    await bot.send_message(message.from_user.id, "Предоставьте мне Excel-файл")
    await EnterForm.waiting_for_reestr.set()


@dp.message_handler(content_types=types.ContentType.DOCUMENT, state=EnterForm.waiting_for_reestr)
async def input_reestr(message: types.Message, state: FSMContext):
    try:
        logging.info("Start input_reestr in BD")
        file_name = message.document.file_name
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        new_file_name = f"reestr_{tomorrow.strftime('%d_%m')}.xlsx"

        logging.info(f"Received file: {file_name}")

        file_id = message.document.file_id
        file_info = await bot.get_file(file_id)
        file_path = file_info.file_path
        downloaded_file = await bot.download_file(file_path)

        current_directory = os.getcwd()
        path_to_save = os.path.join(current_directory, new_file_name)
        logging.info(f"Renaming and saving the file to {path_to_save}")

        with open(path_to_save, "wb") as file:
            file.write(downloaded_file.getvalue())

        logging.info(f"Файл сохранен -  {new_file_name}")

        await message.answer(f"Файл принят и сохранён под именем {new_file_name}")

        ths = await parsed_reestr(new_file_name)

        formatted_message = "Parsed transportation records:\n\n"
        for idx, th in enumerate(ths, start=1):
            formatted_message += f"Transportation {idx}:\n"
            formatted_message += f"Number: {th['num_th']}\n"
            formatted_message += f"Date: {th['date_th']}\n"
            formatted_message += f"Total Count Boxes: {th['total_count_boxes']}\n"
            formatted_message += f"Total Weight: {th['total_weight']}\n\n"

        status, response_message = database.insert_excel_to_db(path_to_save, database.conn)
        await message.answer(response_message)
        await state.finish()

        await message.answer(f"Обработка файла {new_file_name} успешно завершена.")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        await message.answer("Произошла ошибка во время обработки файла.")
        await state.finish()


async def parsed_reestr(excel_file_path):
    ths = []
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = wb.active
    th = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        if row[2] is not None:
            if len(th.keys()) != 0:
                ths.append(th)
            th = {}
            th["num_th"] = row[2]
            if isinstance(row[3], datetime):
                th["date_th"] = row[3].strftime('%Y-%m-%d')
            else:
                th["date_th"] = datetime.strptime(row[3], '%d.%m.%Y').strftime('%Y-%m-%d')
            th["total_count_boxes"] = int(row[9])
            th["total_weight"] = float(row[10])
            th["addresses"] = []
        else:
            addr = {}
            addr["num_route"] = row[4]
            addr["num_shop"] = row[6]
            addr["code_tt"] = row[7]
            addr["address_delivery"] = row[8]
            addr["count_boxes"] = int(row[9])
            addr["weight"] = float(row[10])
            th["addresses"].append(addr)
    ths.append(th)
    return ths


@dp.callback_query_handler(text="Picture")
async def show_daily_report(call: CallbackQuery):
    report_date = datetime.now().date()
    daily_report = get_daily_report(report_date)

    img = Image.new("RGB", (704, 350), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    font = ImageFont.truetype("arial.ttf", 12)

    column_names = ["Номер машины", "Фамилия", "Имя", "Отчество", "Прибытие по регламенту", "Прибытие фактическое", "Погрузка", "Выезд"]
    header_text = "  ".join(column_names)
    draw.text((20, 20), header_text, font=font, fill=(0, 0, 0))

    y = 40
    for row in daily_report:
        car_number, last_name, first_name, patronymic, scheduled_arrival, actual_arrival, shipment, departure = row
        report_text = f"{car_number}  {last_name}  {first_name}  {patronymic} {scheduled_arrival} {actual_arrival} {shipment} {departure}"
        draw.text((20, y), report_text, font=font, fill=(0, 0, 0))
        y += 20

    img.save("daily_report.png")

    with open("daily_report.png", "rb") as f:
        await bot.send_photo(call.from_user.id, f)

    await call.answer()


def date_handler(obj):
    if hasattr(obj, 'isoformat'):  # для объектов date
        return obj.isoformat()
    elif isinstance(obj, Decimal):  # для объектов decimal.Decimal
        return float(obj)
    else:
        raise TypeError("Object of type %s with value of %s is not JSON serializable" % (type(obj), repr(obj)))


from collections import defaultdict


def create_json_from_columns(file_path):
    logging.info('Создание JSON из столбцов начато.')
    data = defaultdict(list)

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        logging.info('Файл Excel успешно открыт.')
    except Exception as e:
        logging.error(f'Ошибка при открытии файла: {e}')
        return

    if not list(sheet.iter_rows(min_row=5, min_col=3, max_col=5, values_only=True)):
        logging.warning('Нет данных в заданном диапазоне столбцов и строк.')

    for idx, row in enumerate(sheet.iter_rows(min_row=5, min_col=3, max_col=5, values_only=True), start=5):
        tn_value = row[0]
        route_value = row[2]
        if tn_value and route_value:
            data[tn_value].append({'num_route': route_value, 'index_number': idx - 4})
            logging.info(f'Добавлено значение: ТН {tn_value}, маршрут {route_value}, индекс {idx - 4}.')
        else:
            logging.warning(f'В строке {idx} отсутствует одно из необходимых значений.')

    if not data:
        logging.warning('Словарь данных пуст после обработки.')

    logging.info('Создание JSON завершено.')
    return {key: value for key, value in data.items()}


def main_json_to_excel():
    logging.info("Start main_json_to_excel")
    try:
        with open('main.txt', 'r', encoding='utf-8') as json_file:
            json_data = json.load(json_file)
        logging.info("Данные JSON успешно загружены")
    except Exception as e:
        logging.error(f"Ошибка при чтении JSON: {e}")
        return

    try:
        #file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
        file_name_today = f"qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today)
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        logging.info("Excel-файл успешно открыт")
    except Exception as e:
        logging.error(f"Ошибка при открытии Excel-файла: {e}")
        return

    try:
        current_th_row = 4  # Начало с четвертой строки
        row_ranges = {}  # Словарь для хранения диапазонов строк для каждого "Номера ТН"

        while current_th_row <= sheet.max_row:
            row = sheet[current_th_row]
            num_th = row[2].value
            if num_th:
                print(num_th)
                th = next((item for item in json_data if item['num_th'] == num_th), None)
                print(th)
                if th:
                    logger.info(f"Обрабатывается num_th: {num_th}, Строка Excel: {row[0].row}")
                    # Вычисляем диапазон строк для текущего "Номера ТН"
                    if num_th not in row_ranges:
                        row_ranges[num_th] = (current_th_row, current_th_row)
                        print(row_ranges[num_th])

                    sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
                    sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
                    sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем для num_th th['num_car']
                    current_row = current_th_row + 1
                    weight = 0.0
                    count_boxes = 0

                    for addr in th['addresses']:
                        sheet.cell(row=current_row, column=5).value = addr['num_route']
                        sheet.cell(row=current_row, column=7).value = addr['num_shop']
                        sheet.cell(row=current_row, column=8).value = addr['code_tt']
                        sheet.cell(row=current_row, column=9).value = addr['address_delivery']
                        sheet.cell(row=current_row, column=10).value = addr['count_boxes']
                        sheet.cell(row=current_row, column=11).value = addr['weight']
                        weight += addr['weight']
                        count_boxes += addr['count_boxes']
                        current_row += 1

                    sheet.cell(row=current_th_row, column=10).value = count_boxes
                    sheet.cell(row=current_th_row, column=11).value = weight
                    current_th_row = current_row
                else:
                    current_th_row += 1
            else:
                current_th_row += 1
        print(row_ranges)
        wb.save('qwerty.xlsx')
        logging.info("Данные успешно записаны в Excel")
    except Exception as e:
        logging.error(f"Ошибка при обработке данных: {e}")

#
# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         current_th_row = 4  # Начало с четвертой строки
#         while current_th_row <= sheet.max_row:
#             row = sheet[current_th_row]
#             num_th = row[2].value
#             if num_th:
#                 th = next((item for item in json_data if item['num_th'] == num_th), None)
#                 if th:
#                     logger.info(f"Обрабатывается num_th: {num_th}, Строка Excel: {row[0].row}")
#
#                     sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
#                     sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
#                     sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем  для num_th  th['num_car']
#                     current_row = current_th_row + 1
#                     weight = 0.0
#                     count_boxes = 0
#
#                     # Создаем список для индексов строк, которые нужно удалить
#                     rows_to_delete = []
#
#                     # Заполняем список индексами строк, которые касаются данного 'th'
#                     while current_row <= sheet.max_row:
#                         num_th_check = sheet.cell(row=current_row, column=3).value
#                         if num_th_check == num_th:
#                             rows_to_delete.append(current_row)
#                         else:
#                             logger.info(f"Прерывание цикла. Не совпало значение num_th_check: {num_th_check} с num_th: {num_th}")
#                             break
#
#                     # Удаляем строки в обратном порядке, чтобы избежать смещения индексов
#                     rows_to_delete.reverse()
#                     for row_to_delete in rows_to_delete:
#                         print(f"Удаление строки {row_to_delete} для num_th: {num_th}")
#                         sheet.delete_rows(row_to_delete)
#
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         weight += addr['weight']
#                         count_boxes += addr['count_boxes']
#                         current_row += 1
#                     sheet.cell(row=current_th_row, column=10).value = count_boxes
#                     sheet.cell(row=current_th_row, column=11).value = weight
#                     current_th_row = current_row
#                 else:
#                     current_th_row += 1
#             else:
#                 current_th_row += 1
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


def remove_extra_rows(sheet, start_row, end_row):
    for row_num in range(end_row, start_row - 1, -1):
        sheet.delete_rows(row_num)


def ranges_addresses_excel_1():
    logging.info("Start ranges_addresses_excel_1")
    try:
        with open('main.txt', 'r', encoding='utf-8') as json_file:
            json_data = json.load(json_file)
        logging.info("Данные JSON успешно загружены")
    except Exception as e:
        logging.error(f"Ошибка при чтении JSON: {e}")
        return
    try:
        while current_th_row <= sheet.max_row:
            row = sheet[current_th_row]
            num_th = row[2].value
            if num_th:
                th = next((item for item in json_data if item['num_th'] == num_th), None)
                if th:
                    logging.info(f"Обработка th={num_th}, строка {current_th_row}")
                    sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
                    sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
                    sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем для num_th th['num_car']
                    current_row = current_th_row + 1
                    weight = 0.0
                    count_boxes = 0
                    for addr in th['addresses']:
                        sheet.cell(row=current_row, column=5).value = addr['num_route']
                        sheet.cell(row=current_row, column=7).value = addr['num_shop']
                        sheet.cell(row=current_row, column=8).value = addr['code_tt']
                        sheet.cell(row=current_row, column=9).value = addr['address_delivery']
                        sheet.cell(row=current_row, column=10).value = addr['count_boxes']
                        sheet.cell(row=current_row, column=11).value = addr['weight']
                        weight += addr['weight']
                        count_boxes += addr['count_boxes']
                        current_row += 1
                    sheet.cell(row=current_th_row, column=10).value = count_boxes
                    sheet.cell(row=current_th_row, column=11).value = weight
                    current_th_row = current_row

                    logging.info(f"Обработка th={num_th} завершена, строка {current_th_row - 1}")
                else:
                    current_th_row += 1
            else:
                current_th_row += 1

    except Exception as e:
        logging.error(f"Ошибка при обработке данных: {e}")


# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#     try:
#         current_th_row = 4  # Начало с четвертой строки
#
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#         while current_th_row <= sheet.max_row:
#             row = sheet[current_th_row]
#             num_th = row[2].value
#             if num_th:
#                 th = next((item for item in json_data if item['num_th'] == num_th), None)
#                 if th:
#                     logging.info(f"Обработка th={num_th}, строка {current_th_row}")
#                     sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
#                     sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
#                     sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем для num_th th['num_car']
#                     current_row = current_th_row + 1
#                     weight = 0.0
#                     count_boxes = 0
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         weight += addr['weight']
#                         count_boxes += addr['count_boxes']
#                         current_row += 1
#                     sheet.cell(row=current_th_row, column=10).value = count_boxes
#                     sheet.cell(row=current_th_row, column=11).value = weight
#                     current_th_row = current_row
#
#                     logging.info(f"Обработка th={num_th} завершена, строка {current_th_row - 1}")
#                 else:
#                     current_th_row += 1
#             else:
#                 current_th_row += 1
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")
#
#

#
# def remove_extra_rows(sheet, start_row, end_row):
#     for row_num in range(end_row, start_row - 1, -1):
#         sheet.delete_rows(row_num)
#
# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#     try:
#         current_th_row = 4  # Начало с четвертой строки
#
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#         while current_th_row <= sheet.max_row:
#             row = sheet[current_th_row]
#             num_th = row[2].value
#             if num_th:
#                 th = next((item for item in json_data if item['num_th'] == num_th), None)
#                 if th:
#                     sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
#                     sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
#                     sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем  для num_th  th['num_car']
#                     current_row = current_th_row + 1
#                     weight = 0.0
#                     count_boxes = 0
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         weight += addr['weight']
#                         count_boxes += addr['count_boxes']
#                         current_row += 1
#                     sheet.cell(row=current_th_row, column=10).value = count_boxes
#                     sheet.cell(row=current_th_row, column=11).value = weight
#                     current_th_row = current_row
#
#                     start_addr_row = current_th_row  # Определите начальную и конечную строку для адресов
#                     end_addr_row = start_addr_row + len(th['addresses']) - 1
#
#                     remove_extra_rows(sheet, start_addr_row, end_addr_row)  # Удаляем лишние строки с адресами
#
#                     if len(th['addresses']) > (current_row - start_addr_row):  # Вставляем новые строки для адресов
#                         for i in range(len(th['addresses']) - (current_row - start_addr_row)):
#                             sheet.insert_rows(current_row)
#                             current_row += 1
#
#                 else:
#                     current_th_row += 1
#             else:
#                 current_th_row += 1
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")
#

# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         current_th_row = 4  # Начало с четвертой строки
#         while current_th_row <= sheet.max_row:
#             row = sheet[current_th_row]
#             num_th = row[2].value
#             if num_th:
#                 th = next((item for item in json_data if item['num_th'] == num_th), None)
#                 if th:
#                     sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
#                     sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
#                     sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем  для num_th  th['num_car']
#                     current_row = current_th_row + 1
#                     weight = 0.0
#                     count_boxes = 0
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         weight += addr['weight']
#                         count_boxes += addr['count_boxes']
#                         current_row += 1
#                     sheet.cell(row=current_th_row, column=10).value = count_boxes
#                     sheet.cell(row=current_th_row, column=11).value = weight
#                     current_th_row = current_row
#
#                     start_addr_row = current_th_row  # Определите начальную и конечную строку для адресов
#                     end_addr_row = start_addr_row + len(th['addresses']) - 1
#
#                     remove_extra_rows(sheet, start_addr_row, end_addr_row)  # Удаляем лишние строки с адресами
#
#                 else:
#                     current_th_row += 1
#             else:
#                 current_th_row += 1
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     row_ranges = {}
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#         current_tn = None
#         start_row = None
#
#         for row in range(5, sheet.max_row + 1):
#             num_th = sheet.cell(row=row, column=3).value
#             first_column_value = sheet.cell(row=row, column=1).value
#             if first_column_value == "Итого":
#                 break
#
#             if num_th and current_tn != num_th:
#                 if current_tn is not None:
#                     row_ranges[current_tn] = (start_row + 1, row - 1)
#                 current_tn = num_th
#                 start_row = row
#
#         if current_tn is not None:
#             row_ranges[current_tn] = (start_row, sheet.max_row if first_column_value != "Итого" else row - 1)
#
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     print(row_ranges)
#
#     try:
#         # for num_th in sorted(row_ranges.keys(), reverse=True):
#         #     start, end = row_ranges[num_th]
#         #     for row_num in range(end, start - 1, -1):
#         #         sheet.delete_rows(row_num)
#         #     print(f"Удалены строки из диапазона {num_th}")
#
#         current_th_row = 4  # Начало с четвертой строки
#         while current_th_row <= sheet.max_row:
#             row = sheet[current_th_row]
#             num_th = row[2].value
#             if num_th:
#                 th = next((item for item in json_data if item['num_th'] == num_th), None)
#                 if th:
#
#                     if num_th in row_ranges:  # Удаление диапазона из row_ranges
#                         del row_ranges[num_th]
#
#                     sheet.cell(row=current_th_row, column=12).value = th['driver']  # Записываем водителя для num_th
#                     sheet.cell(row=current_th_row, column=13).value = "4:00"  # Записываем время для num_th
#                     sheet.cell(row=current_th_row, column=14).value = "CFD3225"  # Записываем  для num_th  th['num_car']
#                     current_row = current_th_row + 1
#                     weight = 0.0
#                     count_boxes = 0
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         weight += addr['weight']
#                         count_boxes += addr['count_boxes']
#                         current_row += 1
#                     sheet.cell(row=current_th_row, column=10).value = count_boxes
#                     sheet.cell(row=current_th_row, column=11).value = weight
#                     current_th_row = current_row
#                 else:
#                     current_th_row += 1
#             else:
#                 current_th_row += 1
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         current_th_row = 4
#         for row in sheet.iter_rows(min_row=current_th_row, values_only=True):
#             if row[2]:
#                 th = next((item for item in json_data if item['num_th'] == row[2]), None)
#                 if th:
#                     if sheet.cell(row=current_th_row, column=3).value == row[2] and not sheet.cell(row=current_th_row,
#                                                                                                    column=12).value:
#                         sheet.cell(row=current_th_row, column=12).value = th['driver']
#                 if th is not None and 'addresses' in th:
#                     current_row = current_th_row + 1
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         sheet.cell(row=current_row, column=12).value = th['driver']
#
#                         current_row += 1
#                     current_th_row = current_row
#                 else:
#                     current_th_row += 1
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


# def main_json_to_excel():
#     logging.info("Start main_json_to_excel")
#     try:
#         with open('main.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         current_th_row = 4
#         for row in sheet.iter_rows(min_row=current_th_row, values_only=True):
#             if row[2]:
#                 th = next((item for item in json_data if item['num_th'] == row[2]), None)
#                 if th is not None and 'addresses' in th:
#                     current_row = current_th_row + 1
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         current_row += 1
#                     current_th_row = current_row
#                 else:
#                     current_th_row += 1
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


# def overweight_json_to_excel():
#     logging.info("Start overweight_json_to_excel")
#     try:
#         with open('overweight.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_modified.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         total_row = None
#         for row in sheet.iter_rows(min_row=4, max_col=1):
#             if row[0].value == "Итого":
#                 total_row = row[0].row
#                 break
#
#         rows_to_move = []
#         for row in sheet.iter_rows(min_row=4, max_row=total_row - 1):
#             num_route = row[4].value
#             if any(item['num_route'] == num_route for item in json_data):
#                 rows_to_move.append((row[0].row, [cell.value for cell in row]))
#
#         for row_index, _ in sorted(rows_to_move, reverse=True):
#             sheet.delete_rows(row_index)
#
#         for _, row_data in rows_to_move:
#             sheet.insert_rows(total_row)
#             for col, value in enumerate(row_data, start=1):
#                 sheet.cell(row=total_row, column=col).value = value
#             total_row += 1
#
#         wb.save('reestr_modified.xlsx')
#         logging.info("overweight_json_to_excel: Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


def ranges_addresses_excel():
    logging.info("Start ranges_addresses_excel")

    try:
        file_name_today = f"qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today)
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        logging.info("Excel-файл успешно открыт")
    except Exception as e:
        logging.error(f"Ошибка при открытии Excel-файла: {e}")
        return

    try:
        row_ranges = {}
        current_tn = None
        start_row = None
        current_row_number = None

        # Вычисление диапазонов строк
        for row in range(5, sheet.max_row + 1):
            num_th = sheet.cell(row=row, column=3).value
            first_column_value = sheet.cell(row=row, column=1).value
            current_row_number = row

            if first_column_value == "Итого":
                break

            if num_th and current_tn != num_th:
                if current_tn is not None:
                    row_ranges[current_tn] = {
                        'start_num_th': start_row,
                        'start_addresses': start_row + 1,
                        'end': row - 1,
                        'count': row - start_row - 1
                    }
                current_tn = num_th
                start_row = row

        if current_tn is not None:
            row_ranges[current_tn] = {
                'start_num_th': start_row,
                'start_addresses': start_row + 1,
                'end': sheet.max_row if first_column_value != "Итого" else row - 1,
                'count': (sheet.max_row if first_column_value != "Итого" else row) - start_row - 1
            }

        if current_row_number is not None:
            print(f"Line number with 'итого': {current_row_number}")


        # Подсчет суммы и запись ее в соответствующую строку
        # for num_th, (start, end) in row_ranges.items():
        #     sum_weight = 0
        #     for row in range(start + 1, end + 1):
        #         try:
        #             weight = float(sheet.cell(row=row, column=11).value)
        #             sum_weight += weight
        #         except (ValueError, TypeError):
        #             pass
        #     sheet.cell(row=start, column=11).value = sum_weight
        #     logging.info(f"Записана сумма веса {sum_weight} для 'Номера ТН' {num_th} в строке {start}")

        wb.save('qwerty.xlsx')
        logging.info("Данные успешно записаны в Excel")
        print(row_ranges)
        return row_ranges, current_row_number

    except Exception as e:
        logging.error(f"Ошибка при обработке данных: {e}")


def ranges_addresses_excel_finish():
    logging.info("Start ranges_addresses_excel")

    try:
        file_name_today = f"qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today)
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        logging.info("Excel-файл успешно открыт")
        with open('main.txt', 'r', encoding='utf-8') as json_file:
            json_data = json.load(json_file)
        logging.info("Данные JSON успешно загружены")
    except Exception as e:
        logging.error(f"Ошибка при открытии Excel-файла: {e}")
        return

    try:
        largest_num_th = max(json_data, key=lambda x: x["num_th"])
        last_code_tt = max(largest_num_th['addresses'], key=lambda x: x['index_number'])['code_tt']
        print(last_code_tt)
        found_row_number = None

        for row in sheet.iter_rows(min_row=50):  # предполагая, что первая строка - это заголовки
            code_tt = row[7].value  # Получение значения из 8-й колонки
            if code_tt == last_code_tt:
                found_row_number = row[0].row
                break

        if found_row_number:
            print(f"Номер строки с {last_code_tt}: {found_row_number}")
        else:
            print(f"Строка с {last_code_tt} не найдена.")
        wb.save('qwerty.xlsx')
        logging.info("Данные успешно записаны в Excel")
        return found_row_number

    except Exception as e:
        logging.error(f"Ошибка при обработке данных: {e}")


# def weight_calculation():
#     logging.info("Start weight_calculation")
#
#     try:
#         file_name_today = f"qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         row_ranges = {}
#         current_tn = None
#         start_row = None
#
#         # Вычисление диапазонов строк
#         for row in range(5, sheet.max_row + 1):
#             num_th = sheet.cell(row=row, column=3).value
#             first_column_value = sheet.cell(row=row, column=1).value
#             if first_column_value == "Итого":
#                 break
#
#             if num_th and current_tn != num_th:
#                 if current_tn is not None:
#                     row_ranges[current_tn] = (start_row + 1, row - 1)
#                 current_tn = num_th
#                 start_row = row
#
#         if current_tn is not None:
#             row_ranges[current_tn] = (start_row + 1, sheet.max_row if first_column_value != "Итого" else row - 1)
#         print(row_ranges)
#
#         # Подсчет суммы и запись ее в соответствующую строку
#         # for num_th, (start, end) in row_ranges.items():
#         #     sum_weight = 0
#         #     for row in range(start + 1, end + 1):
#         #         try:
#         #             weight = float(sheet.cell(row=row, column=11).value)
#         #             sum_weight += weight
#         #         except (ValueError, TypeError):
#         #             pass
#         #     sheet.cell(row=start, column=11).value = sum_weight
#         #     logging.info(f"Записана сумма веса {sum_weight} для 'Номера ТН' {num_th} в строке {start}")
#
#         wb.save('qwerty.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


# def process_json_to_excel():
#     logging.basicConfig(level=logging.INFO)
#     logging.info("Начало обработки данных")
#
#     try:
#         with open('json.txt', 'r', encoding='utf-8') as json_file:
#             json_data = json.load(json_file)
#         logging.info("Данные JSON успешно загружены")
#     except Exception as e:
#         logging.error(f"Ошибка при чтении JSON: {e}")
#         return
#
#     try:
#         file_name_today = f"reestr_{(datetime.now().date() + timedelta(days=1)).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     try:
#         current_th_row = 4
#         for row in sheet.iter_rows(min_row=current_th_row, values_only=True):
#             if row[2]:
#                 th = next((item for item in json_data if item['num_th'] == row[2]), None)
#                 if th is not None and 'addresses' in th:
#                     current_row = current_th_row + 1
#                     for addr in th['addresses']:
#                         sheet.cell(row=current_row, column=5).value = addr['num_route']
#                         sheet.cell(row=current_row, column=7).value = addr['num_shop']
#                         sheet.cell(row=current_row, column=8).value = addr['code_tt']
#                         sheet.cell(row=current_row, column=9).value = addr['address_delivery']
#                         sheet.cell(row=current_row, column=10).value = addr['count_boxes']
#                         sheet.cell(row=current_row, column=11).value = addr['weight']
#                         current_row += 1
#                     current_th_row = current_row
#                 else:
#                     current_th_row += 1
#         wb.save('reestr_modified.xlsx')
#         logging.info("Данные успешно записаны в Excel")
#     except Exception as e:
#         logging.error(f"Ошибка при обработке данных: {e}")


def try_parse_time(time_str):
    try:
        return datetime.strptime(time_str, "%H:%M:%S").time()
    except ValueError:
        return None


@dp.callback_query_handler(text="report14")
async def show_today_report_14(call: CallbackQuery):
    logging.info("Start show_today_report_14")

    data = get_main_json_14()
    print(data)

    today = datetime.now().date()
    file_name_today = f"reestr_{today.strftime('%d_%m')}.xlsx"

    current_directory = os.getcwd()
    path_to_open = os.path.join(current_directory, file_name_today)

    try:
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        current_th_row = 4  # Заголовки находятся на 4-й строке

        rows_to_delete = []
        for row_num, row in enumerate(
                sheet.iter_rows(min_row=current_th_row + 1, max_col=sheet.max_column, max_row=sheet.max_row,
                                values_only=True),
                start=current_th_row + 1):
            logging.info(f"Processing row {row_num}: {row}")
            num_route_from_excel = row[4]  # Пятый столбец с индексом 4
            if num_route_from_excel in data:
                rows_to_delete.append(row_num)

        for row_num in reversed(rows_to_delete):
            try:
                sheet.delete_rows(row_num)
            except Exception as e:
                logging.error(f"Ошибка при удалении строки: {e}")
        print(rows_to_delete)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file_name = temp_file.name
            wb.save(temp_file_name)

        await call.message.answer_document(InputFile(temp_file_name))

    except Exception as e:
        logging.error(f"Ошибка при обработке файла: {e}")


@dp.callback_query_handler(text="report_today")
async def show_today_report(call: CallbackQuery):
    logging.info(f"Start show_today_report")

    data = get_main_json()
    print(data)

    from datetime import datetime

    today = datetime.now().date()
    file_name_today = f"reestr_{today.strftime('%d_%m')}.xlsx"

    current_directory = os.getcwd()
    path_to_open = os.path.join(current_directory, file_name_today)

    try:
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        current_th_row = 4

        new_col_index_arrival = 10
        sheet.insert_cols(new_col_index_arrival)
        sheet.cell(row=4, column=new_col_index_arrival, value="Прибытие")

        new_col_index_departure = 11
        sheet.insert_cols(new_col_index_departure)
        sheet.cell(row=4, column=new_col_index_departure, value="Убытие")

        for row_num, row in enumerate(
                sheet.iter_rows(min_row=current_th_row, max_row=current_th_row + 350, values_only=True),
                start=current_th_row):
            logging.info(f"Processing row {row_num}: {row}")
            num_route_from_excel = row[4]  # Assuming the 5th column has an index of 4
            for addr in data:
                if addr['num_route'] == num_route_from_excel:
                    print(
                        f"Inserting value {addr['arrival_time']} into column {get_column_letter(new_col_index_arrival)}, row {row_num}")
                    try:
                        sheet.cell(row=row_num, column=new_col_index_arrival, value=addr['arrival_time'])
                        sheet.cell(row=row_num, column=new_col_index_arrival).alignment = Alignment(
                            horizontal='left')
                    except Exception as e:
                        logging.error(f"Ошибка при вставке данных: {e}")

                    print(
                        f"Inserting value {addr['departure_time']} into column {get_column_letter(new_col_index_departure)}, row {row_num}")
                    try:
                        sheet.cell(row=row_num, column=new_col_index_departure, value=addr['departure_time'])
                        sheet.cell(row=row_num, column=new_col_index_departure).alignment = Alignment(
                            horizontal='left')
                    except Exception as e:
                        logging.error(f"Ошибка при вставке данных: {e}")

                    break

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file_name = temp_file.name
            wb.save(temp_file_name)

        await call.message.answer_document(InputFile(temp_file_name))
    except Exception as e:
        logging.error(f"Ошибка при обработке файла: {e}")


import json
import openpyxl
from datetime import datetime, timedelta
import os
from aiogram.types import CallbackQuery, InputFile


def reorder_rows_based_on_json(json_data):
    wb = openpyxl.load_workbook('reestr.xlsx')
    sheet = wb.active

    current_order = modify_reestr(json_data)

    for th_data in json_data:
        th_row = current_order[th_data['num_th']]['th_row']

        routes_data = sorted(th_data['addreses'], key=lambda x: x['order'])
        for i, route in enumerate(routes_data):
            current_row_num = current_order[th_data['num_th']]['routes'][route['num_route']]
            new_row_num = th_row + i + 1

            if current_row_num != new_row_num:
                sheet.insert_rows(new_row_num)
                for col_num, cell_value in enumerate(sheet[current_row_num], 1):
                    sheet.cell(row=new_row_num, column=col_num, value=cell_value.value)
                sheet.delete_rows(current_row_num if current_row_num < new_row_num else current_row_num + 1)
    wb.save('reestr.xlsx')


def add_drivers_to_route(main_reestr):
    logging.info("Start add_drivers_to_route")
    drivers_shops = database.get_likes_shop_drivers()
    for th in main_reestr:
        for route in th['addresses']:
            if drivers_shops.get(route['num_shop']):
                th['driver'] = drivers_shops.get(route['num_shop'])
                break
    return main_reestr


# def adjust_rows_for_addresses(addresses_count, row_ranges):
#     offset = 0
#
#     logging.info("Start adjust_rows_for_addresses")
#
#     try:
#         file_name_today = "qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     # Находим строку с "Итого"
#     total_row = None
#     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=1):
#         if row[0].value == "Итого":
#             total_row = row[0].row
#             break
#
#     if total_row is None:
#         logging.error("Строка 'Итого' не найдена")
#         return
#
#     for num_th, count in addresses_count.items():
#         if num_th not in row_ranges:
#             # Копируем данные из последней строки с num_th
#             last_num_th_row = sheet[row_ranges[max(row_ranges, key=row_ranges.get)]['end']]
#             sheet.insert_rows(total_row)
#
#             # Копируем данные
#             for cell in last_num_th_row:
#                 sheet[cell.coordinate.replace(str(cell.row), str(total_row))].value = cell.value
#
#             # Обновляем row_ranges
#             row_ranges[num_th] = {'start': total_row, 'end': total_row, 'count': count}
#             # Обновляем номер строки "Итого"
#             total_row += 1
#
#         # Остальная часть кода без изменений
#         start, end, current_count = row_ranges[num_th]['start'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
#         row_difference = count - current_count
#
#         if row_difference > 0:
#             sheet.insert_rows(end + offset + 1, amount=row_difference)
#             offset += row_difference
#         elif row_difference < 0:
#             for _ in range(abs(row_difference)):
#                 sheet.delete_rows(end + offset)
#                 offset -= 1
#
#         row_ranges[num_th] = {'start': start, 'end': end + offset, 'count': count}
#
#     wb.save('qwerty.xlsx')
#     logging.info("Данные успешно записаны в Excel")
#

def copy_cell_style(src_cell, dest_cell):
    # Копирование стиля, включая шрифт, цвет заливки, границы и другие атрибуты
    dest_cell.font = copy(src_cell.font)
    dest_cell.border = copy(src_cell.border)
    dest_cell.fill = copy(src_cell.fill)
    dest_cell.number_format = src_cell.number_format
    dest_cell.protection = copy(src_cell.protection)
    dest_cell.alignment = copy(src_cell.alignment)


def adjust_rows_for_addresses(addresses_count, row_ranges, finish_row):

    offset = 0
    logging.info("Start adjust_rows_for_addresses")

    try:
        file_name_today = "qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today)
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        logging.info("Excel-файл успешно открыт")

        sheet.delete_rows(finish_row)

    except Exception as e:
        logging.error(f"Ошибка при открытии Excel-файла: {e}")
        return

    for num_th, count in addresses_count.items():
        # Пропускаем num_th, которых нет в row_ranges
        if num_th not in row_ranges:
            continue

    last_key = list(row_ranges.keys())[-1]
    last_start_num_th = row_ranges[last_key]['start_num_th']
    last_start_addresses = row_ranges[last_key]['start_addresses']
    print("last_start_num_th:", last_start_num_th)

    missing_num_th_count = sum(1 for num_th in addresses_count if num_th not in row_ranges)
    additional_rows_to_insert = 2 * missing_num_th_count
    sheet.insert_rows(finish_row, additional_rows_to_insert)
    # Находим отсутствующие num_th в addresses_count
    missing_num_th = [num_th for num_th in addresses_count if num_th not in row_ranges]

    # Выводим список отсутствующих num_th
    print("Отсутствующие num_th:", missing_num_th)

    print(additional_rows_to_insert)
    if missing_num_th_count == 1:
        for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
            dest_cell = sheet.cell(row=finish_row, column=col_num)
            if col_num == 3 and src_cell.value is not None:
                prefix = "0000-0"
                num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
                if num_part.isdigit():  # Увеличиваем число на 1
                    new_value = prefix + str(int(num_part) + 1)
                    dest_cell.value = new_value
                else:
                    dest_cell.value = src_cell.value
            else:
                dest_cell.value = src_cell.value
            copy_cell_style(src_cell, dest_cell)

        for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
            dest_cell = sheet.cell(row=finish_row + 1, column=col_num)
            dest_cell.value = src_cell.value  # Копирование значения ячейки
            copy_cell_style(src_cell, dest_cell)
    else:
        print("Отсутствует номеров ТН:", missing_num_th_count)

    if missing_num_th_count == 2:
        for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
            dest_cell = sheet.cell(row=finish_row, column=col_num)
            if col_num == 3 and src_cell.value is not None:
                prefix = "0000-0"
                num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
                if num_part.isdigit():  # Увеличиваем число на 1
                    new_value = prefix + str(int(num_part) + 1)
                    dest_cell.value = new_value
                else:
                    dest_cell.value = src_cell.value
            else:
                dest_cell.value = src_cell.value
            copy_cell_style(src_cell, dest_cell)

        for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
            dest_cell = sheet.cell(row=finish_row + 1, column=col_num)
            dest_cell.value = src_cell.value  # Копирование значения ячейки
            copy_cell_style(src_cell, dest_cell)

        for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
            dest_cell = sheet.cell(row=finish_row + 2, column=col_num)
            if col_num == 3 and src_cell.value is not None:
                prefix = "0000-0"
                num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
                if num_part.isdigit():  # Увеличиваем число на 2
                    new_value = prefix + str(int(num_part) + 2)
                    dest_cell.value = new_value
                else:
                    dest_cell.value = src_cell.value
            else:
                dest_cell.value = src_cell.value
            copy_cell_style(src_cell, dest_cell)

        for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
            dest_cell = sheet.cell(row=finish_row + 3, column=col_num)
            dest_cell.value = src_cell.value  # Копирование значения ячейки
            copy_cell_style(src_cell, dest_cell)
    else:
        print("Отсутствует номеров ТН:", missing_num_th_count)
    # for num_th, count in addresses_count.items():  # Пропускаем num_th, которых нет в row_ranges
    #     if num_th not in row_ranges:
    #         continue
    #
    #     # Получаем текущий диапазон строк для num_th
    #     start_num_th, start_addresses, end, current_count = row_ranges[num_th]['start_num_th'], row_ranges[num_th]['start_addresses'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
    #     # Вычисляем, сколько строк нужно добавить или удалить
    #     row_difference = count - current_count
    #
    #     if row_difference > 0:
    #         # Добавляем строки, если разница положительная
    #         sheet.insert_rows(end + offset + 1, amount=row_difference)
    #         # Обновляем смещение
    #         offset += row_difference
    #     elif row_difference < 0:
    #         # Удаляем строки, если разница отрицательная
    #         for _ in range(abs(row_difference)):
    #             sheet.delete_rows(end + offset)
    #             # При удалении строк смещение уменьшается
    #             offset -= 1
    #
    #     # Обновляем row_ranges с учетом добавленных или удаленных строк
    #     row_ranges[num_th] = {'start_num_th': start_num_th, 'start_addresses': start_addresses, 'end': end + offset, 'count': count}
    finish_row += additional_rows_to_insert
    print(finish_row)
    wb.save('qwerty.xlsx')
    logging.info("Данные успешно записаны в Excel")


def actually_num_th(addresses_count, row_ranges):

    offset = 0

    logging.info("Start actually_num_th")

    try:
        file_name_today = "qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today)
        wb = openpyxl.load_workbook(path_to_open)
        sheet = wb.active
        logging.info("Excel-файл успешно открыт")
    except Exception as e:
        logging.error(f"Ошибка при открытии Excel-файла: {e}")
        return

    for num_th, count in addresses_count.items():  # Пропускаем num_th, которых нет в row_ranges
        if num_th not in row_ranges:
            continue

        # Получаем текущий диапазон строк для num_th
        start_num_th, start_addresses, end, current_count = row_ranges[num_th]['start_num_th'], row_ranges[num_th]['start_addresses'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
        # Вычисляем, сколько строк нужно добавить или удалить
        row_difference = count - current_count

        if row_difference > 0:
            # Добавляем строки, если разница положительная
            sheet.insert_rows(end + offset + 1, amount=row_difference)
            # Обновляем смещение
            offset += row_difference
        elif row_difference < 0:
            # Удаляем строки, если разница отрицательная
            for _ in range(abs(row_difference)):
                sheet.delete_rows(end + offset)
                # При удалении строк смещение уменьшается
                offset -= 1

        # Обновляем row_ranges с учетом добавленных или удаленных строк
        row_ranges[num_th] = {'start_num_th': start_num_th, 'start_addresses': start_addresses, 'end': end + offset, 'count': count}
    wb.save('qwerty.xlsx')
    logging.info("Данные успешно записаны в Excel")


# def adjust_rows_for_addresses(addresses_count, row_ranges, finish_row):
#
#     offset = 0
#
#     logging.info("Start adjust_rows_for_addresses")
#
#     try:
#         file_name_today = "qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     last_key = list(row_ranges.keys())[-1]
#     last_start_num_th = row_ranges[last_key]['start_num_th']
#     last_start_addresses = row_ranges[last_key]['start_addresses']
#     print(last_start_num_th)
#     sheet.insert_rows(finish_row, 2)  # Вставка 2-х новых строк перед `Итого`
#
#     for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
#         dest_cell = sheet.cell(row=finish_row, column=col_num)
#         if col_num == 3 and src_cell.value is not None:
#             prefix = "0000-0"
#             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#             if num_part.isdigit():  # Увеличиваем число на 1
#                 new_value = prefix + str(int(num_part) + 1)
#                 dest_cell.value = new_value
#             else:
#                 dest_cell.value = src_cell.value
#         else:
#             dest_cell.value = src_cell.value
#         copy_cell_style(src_cell, dest_cell)
#
#     for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
#         dest_cell = sheet.cell(row=finish_row+1, column=col_num)
#         dest_cell.value = src_cell.value  # Копирование значения ячейки
#         copy_cell_style(src_cell, dest_cell)
#
#     for num_th, count in addresses_count.items():  # Пропускаем num_th, которых нет в row_ranges
#         if num_th not in row_ranges:
#             continue
#
#         # Получаем текущий диапазон строк для num_th
#         start_num_th, start_addresses, end, current_count = row_ranges[num_th]['start_num_th'], row_ranges[num_th]['start_addresses'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
#         # Вычисляем, сколько строк нужно добавить или удалить
#         row_difference = count - current_count
#
#         if row_difference > 0:
#             # Добавляем строки, если разница положительная
#             sheet.insert_rows(end + offset + 1, amount=row_difference)
#             # Обновляем смещение
#             offset += row_difference
#         elif row_difference < 0:
#             # Удаляем строки, если разница отрицательная
#             for _ in range(abs(row_difference)):
#                 sheet.delete_rows(end + offset)
#                 # При удалении строк смещение уменьшается
#                 offset -= 1
#
#         # Обновляем row_ranges с учетом добавленных или удаленных строк
#         row_ranges[num_th] = {'start_num_th': start_num_th, 'start_addresses': start_addresses, 'end': end + offset, 'count': count}
#     print(finish_row)
#     wb.save('qwerty.xlsx')
#     logging.info("Данные успешно записаны в Excel")

# def adjust_rows_for_addresses(addresses_count, row_ranges):   Копирует строку
#     """
#     Корректирует количество строк в Excel файле для каждого num_th,
#     чтобы оно соответствовало количеству адресов из addresses_count.
#     """
#
#     offset = 0
#
#     logging.info("Start adjust_rows_for_addresses")
#
#     try:
#         file_name_today = "qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     sheet.insert_rows(125)
#     for col_num, src_cell in enumerate(sheet[121], start=1):
#         dest_cell = sheet.cell(row=125, column=col_num)
#         if col_num == 3 and src_cell.value is not None:
#             prefix = "0000-0"
#             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#             if num_part.isdigit():
#                 # Увеличиваем число на 1
#                 new_value = prefix + str(int(num_part) + 1)
#                 dest_cell.value = new_value
#             else:
#                 dest_cell.value = src_cell.value
#         else:
#             dest_cell.value = src_cell.value
#         copy_cell_style(src_cell, dest_cell)
#     wb.save('qwerty.xlsx')
#     logging.info("Данные успешно записаны в Excel")

# def adjust_rows_for_addresses(addresses_count, row_ranges):
#     """
#     Корректирует количество строк в Excel файле для каждого num_th,
#     чтобы оно соответствовало количеству адресов из addresses_count.
#     """
#
#     offset = 0
#
#     logging.info("Start adjust_rows_for_addresses")
#
#     try:
#         file_name_today = "qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#     sheet.insert_rows(125)
#     for col_num, src_cell in enumerate(sheet[121], start=1):
#         dest_cell = sheet.cell(row=125, column=col_num)
#         if col_num == 3 and src_cell.value is not None:
#             prefix = "0000-0"
#             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#             if num_part.isdigit():
#                 # Увеличиваем число на 1
#                 new_value = prefix + str(int(num_part) + 1)
#                 dest_cell.value = new_value
#             else:
#                 dest_cell.value = src_cell.value
#         else:
#             dest_cell.value = src_cell.value
#         copy_cell_style(src_cell, dest_cell)
#     wb.save('qwerty.xlsx')
#     logging.info("Данные успешно записаны в Excel")

    #
    # total_row = None
    # for row in sheet.iter_rows(min_row=4, max_col=1):
    #     if row[0].value == "Итого":
    #         total_row = row[0].row
    #         break
    #
    # print(total_row)
    # sheet.insert_rows(total_row, amount=2)
    # total_row += 2
    # print(total_row)
    #
    # if total_row is None:
    #     logging.error("Строка 'Итого' не найдена")
    #     return
    #
    # difference = len(addresses_count) - len(row_ranges)
    # print(difference)
    #
    # if difference > 0:
    #     # Находим последний num_th в row_ranges
    #     last_num_th = max(row_ranges, key=lambda x: row_ranges[x]['end'])
    #     start_row = row_ranges[last_num_th]['start']
    #
    #     # Копируем строки start и start - 1
    #     for row_num in range(start_row - 1, start_row + 1):
    #         sheet.insert_rows(sheet.max_row + 1)
    #         for col_num, cell in enumerate(sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=row_num, max_row=row_num), start=1):
    #             sheet.cell(row=sheet.max_row, column=col_num, value=cell[0].value)
    #
    # print(difference)
    #
    # for num_th, count in addresses_count.items():
    #     # Пропускаем num_th, которых нет в row_ranges
    #     if num_th not in row_ranges:
    #         continue
    #
    #     # Получаем текущий диапазон строк для num_th
    #     start, end, current_count = row_ranges[num_th]['start'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
    #     # Вычисляем, сколько строк нужно добавить или удалить
    #     row_difference = count - current_count
    #
    #     if row_difference > 0:
    #         # Добавляем строки, если разница положительная
    #         sheet.insert_rows(end + offset + 1, amount=row_difference)
    #         # Обновляем смещение
    #         offset += row_difference
    #     elif row_difference < 0:
    #         # Удаляем строки, если разница отрицательная
    #         for _ in range(abs(row_difference)):
    #             sheet.delete_rows(end + offset)
    #             # При удалении строк смещение уменьшается
    #             offset -= 1
    #
    #     # Обновляем row_ranges с учетом добавленных или удаленных строк
    #     row_ranges[num_th] = {'start': start, 'end': end + offset, 'count': count}

    # wb.save('qwerty.xlsx')
    # logging.info("Данные успешно записаны в Excel")


def copy_cell_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)


def is_merged_cell(sheet, cell):
    """ Проверка, является ли ячейка частью объединенной области """
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range.coord:
            return True
    return False

def get_start_cell_of_merged_range(sheet, cell):
    """ Получение начальной ячейки объединенной области """
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range.coord:
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    return None


def copy_cell_style_2(src_cell, dest_cell, sheet):
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)


def insert_finish_row(found_row_number):
    logging.info("Start insert_finish_row")

    try:
        file_name_today_1 = f"reestr_{(datetime.now().date()).strftime('%d_%m')}.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today_1)
        wb1 = openpyxl.load_workbook(path_to_open)
        sheet1 = wb1.active
        logging.info("Первый Excel-файл успешно открыт")

        current_row_number = None

        for row in range(5, sheet1.max_row + 1):
            first_column_value = sheet1.cell(row=row, column=1).value
            if first_column_value == "Итого":
                current_row_number = row
                break

        if current_row_number is None:
            logging.error("Строка 'Итого' не найдена")
            return

        # Сохранение данных строки "Итого"
        total_row_data = [sheet1.cell(row=current_row_number, column=col).value for col in range(1, sheet1.max_column + 1)]

    except Exception as e:
        logging.error(f"Ошибка при открытии первого Excel-файла: {e}")
        return

    try:
        file_name_today_2 = "qwerty.xlsx"
        path_to_open = os.path.join(os.getcwd(), file_name_today_2)
        wb2 = openpyxl.load_workbook(path_to_open)
        sheet2 = wb2.active
        logging.info("Второй Excel-файл успешно открыт")

        # Вставка данных в found_row_number + 1
        for col_num, value in enumerate(total_row_data, start=1):
            dest_cell = sheet2.cell(row=found_row_number + 1, column=col_num)
            src_cell = sheet1.cell(row=current_row_number, column=col_num)

            if is_merged_cell(sheet2, dest_cell):
                base_cell = get_start_cell_of_merged_range(sheet2, dest_cell)
                if base_cell:
                    base_cell.value = value
                    copy_cell_style_2(src_cell, base_cell, sheet2)
            else:
                dest_cell.value = value
                copy_cell_style_2(src_cell, dest_cell, sheet2)

        wb2.save(path_to_open)  # Сохранение изменений

    except Exception as e:
        logging.error(f"Ошибка при открытии второго Excel-файла: {e}")
        return

    logging.info("Строка 'Итого' успешно перенесена")


# def insert_finish_row(found_row_number):
#     logging.info("Start insert_finish_row")
#
#     try:
#         file_name_today_1 = f"reestr_{(datetime.now().date()).strftime('%d_%m')}.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today_1)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#         current_row_number = None
#
#         for row in range(5, sheet.max_row + 1):
#             num_th = sheet.cell(row=row, column=3).value
#             first_column_value = sheet.cell(row=row, column=1).value
#             current_row_number = row
#
#             if first_column_value == "Итого":
#                 break
#
#         print(current_row_number)
#
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#
#     try:
#         file_name_today_2 = f"qwerty.xlsx"
#         path_to_open = os.path.join(os.getcwd(), file_name_today_2)
#         wb = openpyxl.load_workbook(path_to_open)
#         sheet = wb.active
#         logging.info("Excel-файл успешно открыт")
#
#         for col_num, src_cell in enumerate(sheet[current_row_number], start=1):
#             dest_cell = sheet.cell(row=found_row_number + 1, column=col_num)
#             dest_cell.value = src_cell.value  # Копирование значения ячейки
#             copy_cell_style(src_cell, dest_cell)
#
#     except Exception as e:
#         logging.error(f"Ошибка при открытии Excel-файла: {e}")
#         return
#
#
#     # Вычисление диапазонов строк
#     for row in range(5, sheet.max_row + 1):
#         num_th = sheet.cell(row=row, column=3).value
#         first_column_value = sheet.cell(row=row, column=1).value
#         current_row_number = row
#
#         if first_column_value == "Итого":
#             break
#
#     print(current_row_number)
#
#     for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
#         dest_cell = sheet.cell(row=finish_row, column=col_num)
#         if col_num == 3 and src_cell.value is not None:
#             prefix = "0000-0"
#             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#             if num_part.isdigit():  # Увеличиваем число на 1
#                 new_value = prefix + str(int(num_part) + 1)
#                 dest_cell.value = new_value
#             else:
#                 dest_cell.value = src_cell.value
#         else:
#             dest_cell.value = src_cell.value
#         copy_cell_style(src_cell, dest_cell)
#
#     for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
#         dest_cell = sheet.cell(row=finish_row + 1, column=col_num)
#         dest_cell.value = src_cell.value  # Копирование значения ячейки
#         copy_cell_style(src_cell, dest_cell)
#
#     # print(additional_rows_to_insert)
#     # if missing_num_th_count == 1:
#     #     for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
#     #         dest_cell = sheet.cell(row=finish_row, column=col_num)
#     #         if col_num == 3 and src_cell.value is not None:
#     #             prefix = "0000-0"
#     #             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#     #             if num_part.isdigit():  # Увеличиваем число на 1
#     #                 new_value = prefix + str(int(num_part) + 1)
#     #                 dest_cell.value = new_value
#     #             else:
#     #                 dest_cell.value = src_cell.value
#     #         else:
#     #             dest_cell.value = src_cell.value
#     #         copy_cell_style(src_cell, dest_cell)
#     #
#     #     for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
#     #         dest_cell = sheet.cell(row=finish_row + 1, column=col_num)
#     #         dest_cell.value = src_cell.value  # Копирование значения ячейки
#     #         copy_cell_style(src_cell, dest_cell)
#     # else:
#     #     print("Отсутствует номеров ТН:", missing_num_th_count)
#     #
#     # if missing_num_th_count == 2:
#     #     for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
#     #         dest_cell = sheet.cell(row=finish_row, column=col_num)
#     #         if col_num == 3 and src_cell.value is not None:
#     #             prefix = "0000-0"
#     #             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#     #             if num_part.isdigit():  # Увеличиваем число на 1
#     #                 new_value = prefix + str(int(num_part) + 1)
#     #                 dest_cell.value = new_value
#     #             else:
#     #                 dest_cell.value = src_cell.value
#     #         else:
#     #             dest_cell.value = src_cell.value
#     #         copy_cell_style(src_cell, dest_cell)
#     #
#     #     for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
#     #         dest_cell = sheet.cell(row=finish_row + 1, column=col_num)
#     #         dest_cell.value = src_cell.value  # Копирование значения ячейки
#     #         copy_cell_style(src_cell, dest_cell)
#     #
#     #     for col_num, src_cell in enumerate(sheet[last_start_num_th], start=1):
#     #         dest_cell = sheet.cell(row=finish_row + 2, column=col_num)
#     #         if col_num == 3 and src_cell.value is not None:
#     #             prefix = "0000-0"
#     #             num_part = src_cell.value[6:]  # Отбрасываем первые шесть символов
#     #             if num_part.isdigit():  # Увеличиваем число на 2
#     #                 new_value = prefix + str(int(num_part) + 2)
#     #                 dest_cell.value = new_value
#     #             else:
#     #                 dest_cell.value = src_cell.value
#     #         else:
#     #             dest_cell.value = src_cell.value
#     #         copy_cell_style(src_cell, dest_cell)
#     #
#     #     for col_num, src_cell in enumerate(sheet[last_start_addresses], start=1):
#     #         dest_cell = sheet.cell(row=finish_row + 3, column=col_num)
#     #         dest_cell.value = src_cell.value  # Копирование значения ячейки
#     #         copy_cell_style(src_cell, dest_cell)
#     # else:
#     #     print("Отсутствует номеров ТН:", missing_num_th_count)
#     # for num_th, count in addresses_count.items():  # Пропускаем num_th, которых нет в row_ranges
#     #     if num_th not in row_ranges:
#     #         continue
#     #
#     #     # Получаем текущий диапазон строк для num_th
#     #     start_num_th, start_addresses, end, current_count = row_ranges[num_th]['start_num_th'], row_ranges[num_th]['start_addresses'], row_ranges[num_th]['end'], row_ranges[num_th]['count']
#     #     # Вычисляем, сколько строк нужно добавить или удалить
#     #     row_difference = count - current_count
#     #
#     #     if row_difference > 0:
#     #         # Добавляем строки, если разница положительная
#     #         sheet.insert_rows(end + offset + 1, amount=row_difference)
#     #         # Обновляем смещение
#     #         offset += row_difference
#     #     elif row_difference < 0:
#     #         # Удаляем строки, если разница отрицательная
#     #         for _ in range(abs(row_difference)):
#     #             sheet.delete_rows(end + offset)
#     #             # При удалении строк смещение уменьшается
#     #             offset -= 1
#     #
#     #     # Обновляем row_ranges с учетом добавленных или удаленных строк
#     #     row_ranges[num_th] = {'start_num_th': start_num_th, 'start_addresses': start_addresses, 'end': end + offset, 'count': count}
#     wb.save(file_name_today)
#     logging.info("Данные успешно записаны в Excel")


@dp.callback_query_handler(text="optimalReport")
async def show_main_report(call: CallbackQuery):
    logging.info("Start show_main_report")
    main_json, addresses_count = get_optimal_json()
    with open('main.txt', 'w', encoding='utf-8') as f:
        json.dump(main_json, f, ensure_ascii=False, indent=4)
    print(main_json)
    addresses_count = {k: addresses_count[k] for k in sorted(addresses_count.keys())}
    print(addresses_count)
    excel_count, finish_row = ranges_addresses_excel()
    adjust_rows_for_addresses(addresses_count, excel_count, finish_row)
    actually_num_th(addresses_count, excel_count)
    main_json_to_excel()
    time.sleep(2)
    found_number = ranges_addresses_excel_finish()
    time.sleep(2)
    insert_finish_row(found_number)


# @dp.callback_query_handler(text="optimalReport")
# async def show_main_report(call: CallbackQuery):
#     logging.info("Start show_main_report")
#     main_json, addresses_count = get_optimal_json()
#     #print("add_drivers_to_route")
#     #print(add_drivers_to_route(main))
#
#     with open('main.txt', 'w', encoding='utf-8') as f:
#         json.dump(main_json, f, ensure_ascii=False, indent=4)
#
#     print(main_json)
#
#     print(addresses_count)
#
#     #main_mapping = main_json_to_excel()
#
#     #print(main_mapping)
#     ranges_addresses_excel()

# @dp.callback_query_handler(text="optimalReport")
# async def show_main_report(call: CallbackQuery):
#     logging.info(f"Start show_main_report")
#     json_data = get_optimal_json()
#
#     with open('json.txt', 'w', encoding='utf-8') as f:
#         json.dump(json_data, f, ensure_ascii=False, indent=4)
#
#     print(json_data)

    # current_mapping = process_json_to_excel()
    #
    # print(current_mapping)


from collections import Counter
import logging


def fix_duplicate_rows(sheet, json_data, current_order):
    logging.info("Starting fix_duplicate_rows...")
    for th_data in json_data:
        routes_data = sorted(th_data['addreses'], key=lambda x: x['order'])
        route_numbers = [route['num_route'] for route in routes_data]

        duplicates = [num for num, count in Counter(route_numbers).items() if count > 1]  # Проверяем наличие дублирующихся строк

        logging.info(f"Found duplicates: {duplicates} for th_data: {th_data['num_th']}")

        if duplicates:
            for duplicate in duplicates:
                missing_route = set(route_numbers) - set(current_order[th_data['num_th']]['routes'].keys())  # Заменяем дублирующую строку на отсутствующую строку

                logging.info(f"Missing routes for duplicate {duplicate}: {missing_route}")

                if missing_route:
                    missing_route = missing_route.pop()
                    current_row_num = current_order[th_data['num_th']]['routes'][duplicate]

                    logging.info(
                        f"Replacing duplicate {duplicate} at row {current_row_num} with missing route {missing_route}")

                    for col_num, cell_value in enumerate(sheet[current_row_num], 1):
                        sheet.cell(row=current_row_num, column=col_num, value=cell_value.value)

                    sheet.cell(row=current_row_num, column=5, value=missing_route)
                    current_order[th_data['num_th']]['routes'][missing_route] = current_row_num


def reorder_addresses_in_excel(json_data, file_name='reestr.xlsx'):
    logging.info("Starting reorder_addresses_in_excel")
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    current_order = get_current_order_from_excel(sheet, json_data)
    original_order = copy.deepcopy(current_order)
    logging.info(f"Current order from Excel: {current_order}")

    fix_duplicate_rows(sheet, json_data, current_order)

    buffered_rows = {}

    for th_data in json_data:
        th_row = current_order[th_data['num_th']]['th_row']
        routes_data = sorted(th_data['addreses'], key=lambda x: x['order'])

        logging.info(f"Starting reordering for th_data: {th_data['num_th']}")

        for i, route in enumerate(routes_data):
            new_row_num = th_row + i + 1

            logging.info(f"Intended placement of route {route['num_route']} in row {new_row_num}")

            current_row_num = original_order[th_data['num_th']]['routes'][route['num_route']]

            if current_row_num != new_row_num:
                logging.info(f"Reordering row from {current_row_num} to {new_row_num} for route: {route['num_route']}")

                buffered_rows[new_row_num] = [(cell.value) for cell in sheet[current_row_num]]

                original_order[th_data['num_th']]['routes'][route['num_route']] = new_row_num

    for new_row_num, row_data in buffered_rows.items():
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=new_row_num, column=col_num, value=cell_value)
        logging.info(f"Updated row {new_row_num} with route {row_data[4]}")

    wb.save(file_name)


def get_current_order_from_excel(sheet, json_data):
    logging.info("Starting get_current_order_from_excel...")
    current_th_row = 4
    result = {}
    for row in sheet.iter_rows(min_row=current_th_row, values_only=True):
        if row[2]:
            th = next((item for item in json_data if item['num_th'] == row[2]), None)
            if th is not None and 'addreses' in th:
                result[th['num_th']] = {
                    "th_row": current_th_row,
                    "routes": {}
                }
                current_row = current_th_row + 1
                while current_row < sheet.max_row + 1:
                    current_addr_row = sheet.cell(row=current_row, column=5).value
                    addr = next((item for item in th['addreses'] if item['num_route'] == current_addr_row), None)
                    if addr:
                        result[th['num_th']]["routes"][addr['num_route']] = current_row
                        current_row += 1
                    else:
                        break
                current_th_row = current_row
            else:
                current_th_row += 1
    return result


@dp.callback_query_handler(text="reportForCustomer")
async def start_reportForCustomer(query: types.CallbackQuery):
    await bot.send_message(
        query.from_user.id, "За какой день Вам нужен отчёт ?",
        reply_markup=nav.mainMenu3)


@dp.callback_query_handler(text="internalReport")
async def show_daily_report_excel(call: CallbackQuery):

    logging.info("start show_daily_report_excel")

    temp_file_path = "report_internal.xlsx"

    workbook = Workbook()
    sheet = workbook.active

    current_date = datetime.now().date()

    header_data = [
        ('Погрузка и доставка продукции по ТТ',),
        (f'Дата: {current_date}',),
        ('Номер авто', 'Водитель', 'Время\n прибытия', 'Время начала\nпогрузки', 'Время\nвыезда',
         'Время прибытия\nпо регламенту', 'Опоздание', 'Простой', 'Время\nпогрузки', 'Цель\nвъезда/выезда')
    ]

    info_for_report = get_info_for_report()

    report_data = [
        (row[2],  # num_car
         row[1],  # lastname
         row[3].strftime('%H:%M') if row[3] else '',  # arrival_time_fact
         row[4].strftime('%H:%M') if row[4] else '',  # shipment_time
         row[5].strftime('%H:%M') if row[5] else '',  # departure_time
         row[6].strftime('%H:%M') if row[6] else '',  # arrival_time
         str((datetime.combine(datetime.min, row[3]) - datetime.combine(
                 datetime.min, row[6])).seconds // 60) + ' мин' if row[3] and row[6] and row[3] > row[6] else '',  # Опоздание
         str((datetime.combine(datetime.min, row[6]) - datetime.combine(
             datetime.min, row[4])).seconds // 60 * -1) + ' мин' if row[6] and row[4] and row[4] < row[6] else (
             str((datetime.combine(datetime.min, row[4]) - datetime.combine(
                 datetime.min, row[6])).seconds // 60) + ' мин' if row[6] and row[4] else ''),  # Простой
         str((datetime.combine(datetime.min, row[5]) - datetime.combine(
             datetime.min, row[4])).seconds // 60) + ' мин' if row[4] and row[5] else '',  # Время погрузки
         'погрузка'  # Цель въезда/выезда
         ) for row in info_for_report if row[1]]

    data = header_data + report_data

    for row in data:
        sheet.append(row)

    for row_idx in [1, 2]:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
        cell = sheet.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal='center')

    for cell in sheet[3]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sheet.row_dimensions[3].height = 55

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        sheet.column_dimensions[col_letter].width = 18

    workbook.save(temp_file_path)
    with open(temp_file_path, "rb") as excel_file:
        await bot.send_document(call.from_user.id, InputFile(excel_file, filename="report_internal.xlsx"))
    os.remove(temp_file_path)
    await call.answer()


if __name__ == "__main__":
    executor.start_polling(dp, skip_updates=True)
